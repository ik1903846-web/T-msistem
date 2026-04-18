import streamlit as st
import pandas as pd
import json
import pickle
import io
from datetime import datetime

from unified_engine import (UnifiedEngine, DerinAnaliz, read_excel_bytes, donem_from_filename,
                             fmt_milyon, safe_float, C_ROE, hesapla_pd, roe_istikrar_hesapla, roe_donus_hesapla)

st.set_page_config(
    page_title="BIST Analiz Sistemi",
    page_icon="\U0001f4ca",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&display=swap');

* { font-family: 'Space Grotesk', sans-serif; }

[data-testid="stAppViewContainer"] { background:#080E17; }
[data-testid="stSidebar"] { background:#060D18; border-right:1px solid #0F2040; }
[data-testid="stSidebar"] * { color:#94A3B8 !important; }
section.main > div { padding-top:.5rem; }
h1,h2,h3 { color:#E2E8F0 !important; }
p, li { color:#94A3B8 !important; }

/* ── Sidebar ── */
.sb-brand { font-size:22px; font-weight:800; letter-spacing:-0.5px; padding:12px 0 2px 0; }
.sb-brand span { color:#38BDF8; }
.sb-sub { font-size:10px; color:#1E3448 !important; letter-spacing:2px;
  text-transform:uppercase; margin-bottom:16px; }
.sb-card { background:#0D1926; border:1px solid #0F2040; border-radius:10px;
  padding:12px 14px; margin:6px 0; }
.sb-card-label { font-size:9px; color:#1E3448 !important; text-transform:uppercase;
  letter-spacing:1.5px; margin-bottom:4px; }
.sb-card-val { font-size:18px; font-weight:800; color:#E2E8F0 !important; }
.sb-card-sub { font-size:10px; color:#1E3448 !important; margin-top:2px; }

/* ── Page header ── */
.ph { padding:22px 28px 18px; margin-bottom:20px;
  border-bottom:1px solid #0F2040; }
.ph-badge { display:inline-block; font-size:9px; font-weight:700;
  letter-spacing:2px; text-transform:uppercase; padding:3px 10px;
  border-radius:20px; margin-bottom:8px; }
.ph-title { font-size:24px; font-weight:800; color:#E2E8F0; margin:0; letter-spacing:-0.5px; }
.ph-sub { font-size:12px; color:#475569; margin-top:4px; }

/* ── Metrik kartlar ── */
.mrow { display:flex; gap:10px; margin-bottom:20px; flex-wrap:wrap; }
.mc { flex:1; min-width:90px; background:#0D1926; border:1px solid #0F2040;
  border-radius:12px; padding:14px 16px; position:relative; overflow:hidden; }
.mc::before { content:''; position:absolute; top:0; left:0; right:0; height:2px; }
.mc-green::before { background:linear-gradient(90deg,#4ADE80,#22D3EE); }
.mc-yellow::before { background:linear-gradient(90deg,#FCD34D,#F59E0B); }
.mc-blue::before { background:linear-gradient(90deg,#38BDF8,#6366F1); }
.mc-purple::before { background:linear-gradient(90deg,#A78BFA,#EC4899); }
.mc-red::before { background:linear-gradient(90deg,#F87171,#FB923C); }
.mc-num { font-size:28px; font-weight:900; line-height:1; }
.mc-lbl { font-size:9px; color:#475569; margin-top:4px;
  text-transform:uppercase; letter-spacing:1px; }

/* ── Badge ── */
.b-guclu { background:#14532D;color:#4ADE80;padding:2px 9px;border-radius:20px;
  font-size:10px;font-weight:700;border:1px solid #166534; }
.b-pot { background:#422006;color:#FCD34D;padding:2px 9px;border-radius:20px;
  font-size:10px;font-weight:700;border:1px solid #92400E; }
.b-zayif { background:#431407;color:#FB923C;padding:2px 9px;border-radius:20px;
  font-size:10px;font-weight:700;border:1px solid #9A3412; }
.b-elen { background:#450A0A;color:#F87171;padding:2px 9px;border-radius:20px;
  font-size:10px;font-weight:700;border:1px solid #991B1B; }

/* ── Kesisim kart ── */
.kk { background:#0D1926; border:1px solid #0F2040; border-radius:12px;
  padding:16px 18px; margin-bottom:10px; transition:all .15s; }
.kk:hover { border-color:#1E3448; background:#111F30; }
.kk-kod { font-size:18px; font-weight:900; color:#38BDF8; }
.kk-sektor { font-size:11px; color:#475569; }
.kk-scores { display:flex; gap:8px; margin:10px 0 8px; }
.kk-score { background:#080E17; border-radius:8px; padding:8px 12px; flex:1; text-align:center; }
.kk-score-val { font-size:20px; font-weight:900; }
.kk-score-lbl { font-size:9px; color:#475569; text-transform:uppercase; letter-spacing:1px; margin-top:2px; }
.kk-metrics { display:flex; gap:6px; flex-wrap:wrap; }
.kk-chip { background:#0F2040; border:1px solid #1E3448; border-radius:6px;
  padding:4px 10px; font-size:11px; color:#94A3B8; }
.kk-chip b { color:#E2E8F0; }

/* ── Upload area ── */
[data-testid="stFileUploader"] { background:#0D1926; border-radius:10px;
  border:1px dashed #1E3448; }

/* ── Buttons ── */
[data-testid="stButton"] button { background:#0F2040; color:#94A3B8;
  border:1px solid #1E3448; border-radius:8px; font-size:12px; }
button[kind="primary"] { background:linear-gradient(135deg,#1E40AF,#1E3A8A) !important;
  border:none !important; color:white !important; }

/* ── Inputs ── */
[data-testid="stMultiSelect"] > div,
[data-testid="stSelectbox"] > div > div { background:#0D1926 !important;
  border:1px solid #1E3448 !important; border-radius:8px !important; }

/* ── DataFrame ── */
[data-testid="stDataFrame"] { background:#0D1926; border-radius:10px; overflow:hidden; }
[data-testid="stDataFrame"] table { font-size:12px !important; }
[data-testid="stDataFrame"] th { background:#060D18 !important; color:#475569 !important;
  font-weight:600 !important; }
[data-testid="stDataFrame"] td { color:#CBD5E1 !important; }
[data-testid="stDataFrame"] tr:hover td { background:#111F30 !important; }

/* ── Expander ── */
[data-testid="stExpander"] { background:#0D1926; border:1px solid #0F2040; border-radius:10px; }
[data-testid="stExpander"] summary { color:#475569 !important; }

/* ── Tabs ── */
[data-testid="stTabs"] [role="tab"] { color:#475569 !important; font-size:12px; padding:8px 16px; }
[data-testid="stTabs"] [role="tab"][aria-selected="true"] { color:#38BDF8 !important;
  border-bottom:2px solid #38BDF8 !important; }

hr { border-color:#0F2040 !important; }
[data-testid="stAlert"] { border-radius:10px; }
</style>""", unsafe_allow_html=True)

# ── SESSION STATE ─────────────────────────────────────────────────────────────
for k,v in [('quarters',{}),('engine',None),('son_donem',None),('son_yukleme',None),
             ('watchlist',{}),('geri_yil',3)]:
    if k not in st.session_state: st.session_state[k]=v

def df_to_excel_bytes(df):
    import io
    import openpyxl
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    for ci, col in enumerate(df.columns, 1):
        ws.cell(1, ci, str(col))
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(ri, ci, val if val is not None else '')
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def donem_fmt(d):
    return f"{d[:4]}/{d[4:]}" if d and len(d)==6 else (d or '-')

def badge(k):
    cls = {'GUCLU ADAY':'b-guclu','POTANSIYEL':'b-pot','ZAYIF':'b-zayif','ELENDI':'b-elen'}.get(k,'b-elen')
    lbl = {'GUCLU ADAY':'\U0001f7e2 GUCLU','POTANSIYEL':'\U0001f7e1 POTANSIYEL',
           'ZAYIF':'\U0001f7e0 ZAYIF','ELENDI':'\U0001f534 ELENDI'}.get(k,k)
    return f"<span class='{cls}'>{lbl}</span>"

KARAR_RENK = {'GUCLU ADAY':'#4ADE80','POTANSIYEL':'#FCD34D','ZAYIF':'#FB923C','ELENDI':'#F87171'}

# ── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("<div class='sb-brand'>BIST <span>ANALiZ</span></div>", unsafe_allow_html=True)
    st.markdown("<div class='sb-sub'>FARK · GERI · BEBEK · KESISIM</div>", unsafe_allow_html=True)

    page = st.radio("", [
        "\U0001f50d FARK Scanner",
        "\U0001f4c9 GER\u0130 Taray\u0131c\u0131",
        "\U0001f3af Kesisim",
        "\U0001f476 Bebek Hisse",
        "\U0001f4ca Detay Analizi",
        "\U0001f504 ROE Tarayici",
        "\u2b50 Takip Listesi",
        "\U0001f4da Metodoloji",
        "\u2699\ufe0f Ayarlar"
    ], label_visibility="collapsed")

    st.markdown("<hr>", unsafe_allow_html=True)

    # Periyot (GERİ için)
    st.markdown("<div class='sb-card-label'>Buyume Periyotu</div>", unsafe_allow_html=True)
    yil = st.radio("", [1,2,3,4,5], horizontal=True,
                    index=[1,2,3,4,5].index(st.session_state.geri_yil),
                    label_visibility="collapsed",
                    format_func=lambda x: f"{x}Y")
    st.session_state.geri_yil = yil
    st.markdown("<hr>", unsafe_allow_html=True)

    engine = st.session_state.engine
    if engine and st.session_state.son_yukleme:
        donems = sorted(st.session_state.quarters.keys())
        gun = (datetime.now()-datetime.fromisoformat(st.session_state.son_yukleme)).days
        renk = '#EF4444' if gun>85 else '#4ADE80'
        istat = engine.istatistik(yil)
        st.markdown(f"""
        <div class='sb-card'>
          <div class='sb-card-label'>Veri</div>
          <div class='sb-card-val' style='color:{renk}'>{len(donems)} donem</div>
          <div class='sb-card-sub'>Son: {donem_fmt(donems[-1])} · {gun}g once</div>
        </div>
        <div class='sb-card'>
          <div class='sb-card-label'>Tarama Sonucu</div>
          <div style='display:flex;gap:8px;margin-top:4px'>
            <div style='flex:1;text-align:center'>
              <div style='font-size:20px;font-weight:900;color:#38BDF8'>{istat['fark']}</div>
              <div style='font-size:9px;color:#475569'>FARK</div>
            </div>
            <div style='flex:1;text-align:center'>
              <div style='font-size:20px;font-weight:900;color:#A78BFA'>{istat['geri']}</div>
              <div style='font-size:9px;color:#475569'>GER\u0130</div>
            </div>
            <div style='flex:1;text-align:center'>
              <div style='font-size:20px;font-weight:900;color:#4ADE80'>{istat['kesisim']}</div>
              <div style='font-size:9px;color:#475569'>KES\u0130S\u0130M</div>
            </div>
          </div>
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown("""<div class='sb-card'>
          <div class='sb-card-label'>Veri</div>
          <div style='color:#1E3448;font-size:12px'>Yuklenmedi</div>
        </div>""", unsafe_allow_html=True)

    if st.session_state.watchlist:
        st.markdown(f"<div style='color:#FCD34D;font-size:12px;margin-top:8px'>"
                    f"\u2b50 {len(st.session_state.watchlist)} takipte</div>",
                    unsafe_allow_html=True)

    st.markdown("<div style='margin-top:20px;font-size:9px;color:#0F2040'>"
                "v2.0 · GXSMODUJ Metodolojisi</div>", unsafe_allow_html=True)

# ── VERİ YÜKLEME COMPONENTI ──────────────────────────────────────────────────
def veri_yukle_widget():
    with st.expander("\U0001f4c1 Veri Yukle", expanded=not bool(st.session_state.engine)):

        tab_parquet, tab_excel = st.tabs(["\U0001f4be Kayitli Veri Yukle (.parquet)", "\U0001f4ca Excel Dosyalari Yukle"])

        # ── Parquet yükleme ──────────────────────────────────────────────────
        with tab_parquet:
            st.markdown("""<div style='background:#0A1C0F;border:1px solid #166534;
            border-radius:8px;padding:10px 14px;font-size:12px;color:#4ADE80;margin-bottom:10px'>
            \u26a1 En hizli yontem — daha once kaydedilmis .parquet dosyasini yukle
            </div>""", unsafe_allow_html=True)
            pq_file = st.file_uploader("", type=["parquet"], label_visibility="collapsed",
                                        key="pq_uploader")
            if pq_file:
                if st.button("\u26a1 Yukle", type="primary", use_container_width=True, key="pq_btn"):
                    with st.spinner("Yukleniyor..."):
                        try:
                            meta_df = pd.read_parquet(io.BytesIO(pq_file.read()), engine="pyarrow")
                            quarters = {}
                            for (donem, kod), row in meta_df.iterrows():
                                quarters.setdefault(donem, {})[kod] = row.to_dict()
                            eng = UnifiedEngine(quarters)
                            son_yuk = meta_df.attrs.get("son_yukleme", datetime.now().isoformat())
                            st.session_state.update({
                                "quarters": quarters, "engine": eng,
                                "son_donem": eng.son_donem, "son_yukleme": son_yuk
                            })
                            st.success(f"\u2713 {len(quarters)} donem · {len(eng.son_data)} hisse")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Parquet yuklenemedi: {e}")

        # ── Excel yükleme ────────────────────────────────────────────────────
        with tab_excel:
            st.markdown("""<div style='background:#0A1628;border:1px solid #1E3448;
            border-radius:8px;padding:10px 14px;font-size:12px;color:#64748B;margin-bottom:10px'>
            Fastweb → Sirket Puanlama → Model: <b style='color:#94A3B8'>Uygulama</b> →
            Sektör: <b style='color:#94A3B8'>Tumu</b> →
            Donem: <b style='color:#4ADE80'>Spesifik sec (Cari Donem degil!)</b>
            </div>""", unsafe_allow_html=True)
            if "xlsx_key" not in st.session_state:
                st.session_state.xlsx_key = 0

            uploaded = st.file_uploader("", type=['xlsx'], accept_multiple_files=True,
                                         label_visibility="collapsed",
                                         key=f"xlsx_{st.session_state.xlsx_key}")

            if uploaded:
                c1,c2,c3 = st.columns([2,1,1])
                with c1:
                    st.markdown(f"<p style='color:#38BDF8;font-size:13px;margin-top:8px'>"
                                f"\U0001f4ce {len(uploaded)} dosya secildi</p>", unsafe_allow_html=True)
                with c2:
                    if st.button("\U0001f680 Taramayi Baslat", type="primary", use_container_width=True, key="baslat_btn_v2"):
                        with st.spinner("Analiz ediliyor..."):
                            quarters, hatalar = {}, []
                            # Mevcut quarters'i koru (ek donem ekleme)
                            quarters = dict(st.session_state.quarters) if st.session_state.quarters else {}
                            for f in uploaded:
                                d = donem_from_filename(f.name)
                                if d:
                                    data = read_excel_bytes(f.read())
                                    if data: quarters[d] = data
                                    else: hatalar.append(f"`{f.name}` okunamadi")
                                else:
                                    hatalar.append(f"`{f.name}` donem cikarilamadi")
                            if quarters:
                                eng = UnifiedEngine(quarters)
                                st.session_state.update({
                                    "quarters": quarters, "engine": eng,
                                    "son_donem": eng.son_donem,
                                    "son_yukleme": datetime.now().isoformat()
                                })
                                st.success(f"\u2713 {len(quarters)} donem · {len(eng.son_data)} hisse")
                                if hatalar:
                                    for h in hatalar: st.warning(h)
                                st.rerun()
                            else:
                                st.error("Hicbir dosya yuklenemedi")
                                for h in hatalar: st.warning(h)
                with c3:
                    if st.button("\U0001f5d1\ufe0f Temizle", use_container_width=True, key="temizle_btn_v2",
                                  help="Secili tum dosyalari kaldir"):
                        st.session_state.xlsx_key += 1
                        st.rerun()

        # ── Kaydet butonu (veri varsa göster) ───────────────────────────────
        if st.session_state.engine:
            st.markdown("<hr>", unsafe_allow_html=True)
            donems = sorted(st.session_state.quarters.keys())
            st.markdown(f"""<div style='background:#0D1926;border:1px solid #0F2040;
            border-radius:8px;padding:10px 14px;font-size:12px;margin-bottom:8px'>
            <span style='color:#475569'>{len(donems)} donem bellekte</span>
            <span style='color:#1E3448'> · </span>
            <span style='color:#475569'>{donem_fmt(donems[0])} — {donem_fmt(donems[-1])}</span>
            </div>""", unsafe_allow_html=True)
            if st.button("\U0001f4be Veriyi .parquet Olarak Kaydet", use_container_width=True):
                try:
                    rows = []
                    for donem, hisseler in st.session_state.quarters.items():
                        for kod, veriler in hisseler.items():
                            rows.append({"donem": donem, "kod": kod, **veriler})
                    df_save = pd.DataFrame(rows).set_index(["donem","kod"])
                    df_save.attrs["son_yukleme"] = st.session_state.son_yukleme or datetime.now().isoformat()
                    buf = io.BytesIO()
                    df_save.to_parquet(buf, engine="pyarrow")
                    buf.seek(0)
                    st.download_button(
                        "\u2b07\ufe0f bist_veri.parquet indir",
                        data=buf.getvalue(),
                        file_name="bist_veri.parquet",
                        mime="application/octet-stream",
                        use_container_width=True
                    )
                    st.success("Hazir! Dosyayi bilgisayarinda sakla, bir daha Excel yuklemene gerek yok.")
                except Exception as e:
                    st.error(f"Kayit hatasi: {e}")

def bos_ekran(emoji, mesaj):
    st.markdown(f"""<div style='background:#0D1926;border:1px dashed #1E3448;
    border-radius:14px;padding:80px;text-align:center;margin-top:20px'>
    <div style='font-size:56px'>{emoji}</div>
    <p style='color:#475569;margin-top:12px;font-size:14px'>{mesaj}</p>
    </div>""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 1: FARK SCANNER
# ════════════════════════════════════════════════════════════════════════════
if page == "\U0001f50d FARK Scanner":
    st.markdown("""<div class='ph'>
    <div class='ph-badge' style='background:#0A1628;color:#38BDF8;border:1px solid #1E3448'>
    FARK SISTEMI</div>
    <div class='ph-title'>Fiyat Ardinda Kalan Sirketler</div>
    <div class='ph-sub'>Operasyonel buyume gosterirken piyasanin gerisinde kalan hisseleri tespit et</div>
    </div>""", unsafe_allow_html=True)

    veri_yukle_widget()
    engine = st.session_state.engine
    if not engine:
        bos_ekran('\U0001f50d','Veri yukleyerek FARK taramasini baslat'); st.stop()

    sonuclar = engine.fark_tara()
    guclu = [r for r in sonuclar if r['puan']>=75]
    pot   = [r for r in sonuclar if 55<=r['puan']<75]
    zayif = [r for r in sonuclar if 35<=r['puan']<55]

    st.markdown(f"""<div class='mrow'>
    <div class='mc mc-green'><div class='mc-num' style='color:#4ADE80'>{len(guclu)}</div>
      <div class='mc-lbl'>Guclu Aday</div></div>
    <div class='mc mc-yellow'><div class='mc-num' style='color:#FCD34D'>{len(pot)}</div>
      <div class='mc-lbl'>Potansiyel</div></div>
    <div class='mc mc-red'><div class='mc-num' style='color:#FB923C'>{len(zayif)}</div>
      <div class='mc-lbl'>Zayif</div></div>
    <div class='mc mc-blue'><div class='mc-num' style='color:#38BDF8'>{len(sonuclar)}</div>
      <div class='mc-lbl'>Toplam Gecti</div></div>
    <div class='mc mc-red'><div class='mc-num' style='color:#F87171'>{len(engine.son_data)-len(sonuclar)}</div>
      <div class='mc-lbl'>Elendi</div></div>
    </div>""", unsafe_allow_html=True)

    arama_col2, bos2 = st.columns([1.5, 4.5])
    with arama_col2:
        arama_fark = st.text_input("\U0001f50d Hisse Ara", placeholder="Kod ara...",
                                    label_visibility="collapsed", key="fark_ara_v2")

    c1,c2,c3,c4 = st.columns([1.5,2,1.5,1])
    with c1:
        kf = st.multiselect("Karar",["GUCLU ADAY","POTANSIYEL","ZAYIF","ELENDI"],
                             default=["GUCLU ADAY","POTANSIYEL"])
    with c2:
        sf = st.multiselect("Sektor", sorted(set(r['sektor'] for r in sonuclar if r['sektor'])))
    with c3:
        sir = st.selectbox("Sirala",["FARK Puani \u2193","FK/PD% \u2193","Buyume% \u2193"])
    with c4:
        mp = st.number_input("Min Puan",0,100,0,5)

    goster = [r for r in sonuclar if r['karar'] in kf and r['puan']>=mp]
    if arama_fark: goster = [r for r in goster if arama_fark.upper() in r['kod'].upper()]
    if sf: goster = [r for r in goster if r['sektor'] in sf]
    if sir == "FK/PD% \u2193": goster.sort(key=lambda r: r.get('fkpd') or 0, reverse=True)
    elif sir == "Buyume% \u2193": goster.sort(key=lambda r: r.get('buyume') or 0, reverse=True)

    st.markdown(f"<p style='font-size:11px;color:#475569;margin:8px 0'>"
                f"{len(goster)} hisse · Donem: <b style='color:#94A3B8'>"
                f"{donem_fmt(st.session_state.son_donem)}</b></p>", unsafe_allow_html=True)

    tablo = pd.DataFrame([{
        '\u2b50':       r['kod'] in st.session_state.watchlist,
        'Kod':          r['kod'],
        'Sektor':       r['sektor'],
        'Puan':         int(r['puan']),
        'Karar':        r['karar'],
        'A':            r['A'], 'B': r['B'], 'C': r['C'], 'D': r['D'],
        'Faal.Kari':    fmt_milyon(r.get('fk')),
        'Piy.Degeri':   fmt_milyon(r.get('pd')),
        'PD/DD':        round(r['pddd'],2) if r.get('pddd') else None,
        'FK/PD%':       round(r['fkpd'],1) if r.get('fkpd') else None,
        'Marj%':        round(r['marj'],1) if r.get('marj') else None,
        'Buyume%':      round(r.get('buyume',0),0) if r.get('buyume') is not None else None,
        'ROE 30+D':     roe_istikrar_hesapla(engine.quarters, engine.sorted_donems, r['kod'])[0],
    } for r in goster])

    edited = st.data_editor(tablo, column_config={
        '\u2b50':       st.column_config.CheckboxColumn('\u2b50',width='small'),
        'Kod':          st.column_config.TextColumn('Kod',width='small'),
        'Sektor':       st.column_config.TextColumn('Sektor',width='medium'),
        'Puan':         st.column_config.NumberColumn('Puan',width='small',format='%d'),
        'Karar':        st.column_config.TextColumn('Karar',width='medium'),
        'A':            st.column_config.NumberColumn('A',width='small'),
        'B':            st.column_config.NumberColumn('B',width='small'),
        'C':            st.column_config.NumberColumn('C',width='small'),
        'D':            st.column_config.NumberColumn('D',width='small'),
        'Faal.Kari':    st.column_config.TextColumn('Faal.Kari',width='small'),
        'Piy.Degeri':   st.column_config.TextColumn('Piy.Degeri',width='small'),
        'PD/DD':        st.column_config.NumberColumn('PD/DD',width='small',format='%.2f'),
        'FK/PD%':       st.column_config.NumberColumn('FK/PD%',width='small',format='%.1f'),
        'Marj%':        st.column_config.NumberColumn('Marj%',width='small',format='%.1f'),
        'Buyume%':      st.column_config.NumberColumn('Buyume%',width='small',format='%.0f'),
    }, disabled=[c for c in tablo.columns if c!='\u2b50'],
    hide_index=True, use_container_width=True,
    height=min(40+len(goster)*35,600), key='fark_tbl')

    for i,row in edited.iterrows():
        kod,istek = row['Kod'], row['\u2b50']
        in_wl = kod in st.session_state.watchlist
        if istek and not in_wl:
            r = goster[i]
            st.session_state.watchlist[kod] = {
                'puan':r['puan'],'karar':r['karar'],'sektor':r['sektor'],
                'sistem':'FARK','eklenme':datetime.now().strftime('%Y-%m-%d'),
                'donem':st.session_state.son_donem}
            st.toast(f"\u2b50 {kod} eklendi!",icon="\u2705"); st.rerun()
        elif not istek and in_wl:
            del st.session_state.watchlist[kod]; st.rerun()

    # Excel indir
    if goster:
        xls_df = tablo.drop(columns=['⭐'], errors='ignore')
        st.download_button(
            "⬇️ Listeyi Excel Olarak indir",
            data=df_to_excel_bytes(xls_df),
            file_name=f"FARK_{st.session_state.son_donem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 2: GERI TARAYICI
# ════════════════════════════════════════════════════════════════════════════
elif page == "\U0001f4c9 GER\u0130 Taray\u0131c\u0131":
    st.markdown(f"""<div class='ph'>
    <div class='ph-badge' style='background:#0D0A28;color:#A78BFA;border:1px solid #2D1F6E'>
    GER\u0130 S\u0130STEM\u0130</div>
    <div class='ph-title'>Fiyat Gerisinde Kalan Sirketler</div>
    <div class='ph-sub'>F/K &lt; 30 · PD/DD &lt; 5 · EFK buyumesi fiyatin onunde olan hisseleri tespit et · {yil}Y periyot</div>
    </div>""", unsafe_allow_html=True)

    veri_yukle_widget()
    engine = st.session_state.engine
    if not engine:
        bos_ekran('\U0001f4c9','Veri yukleyerek GERI taramasini baslat'); st.stop()

    sonuclar = engine.geri_tara(yil)
    guclu  = [r for r in sonuclar if r['puan']>=75]
    pot    = [r for r in sonuclar if 55<=r['puan']<75]
    geride = [r for r in sonuclar if r.get('fiyat_geride')]
    elen   = len(engine.son_data) - len(sonuclar)

    st.markdown(f"""<div class='mrow'>
    <div class='mc mc-green'><div class='mc-num' style='color:#4ADE80'>{len(guclu)}</div>
      <div class='mc-lbl'>Guclu Aday</div></div>
    <div class='mc mc-yellow'><div class='mc-num' style='color:#FCD34D'>{len(pot)}</div>
      <div class='mc-lbl'>Potansiyel</div></div>
    <div class='mc mc-blue'><div class='mc-num' style='color:#38BDF8'>{len(sonuclar)}</div>
      <div class='mc-lbl'>Filtre Gecti</div></div>
    <div class='mc mc-green'><div class='mc-num' style='color:#4ADE80'>{len(geride)}</div>
      <div class='mc-lbl'>Fiyat Geride</div></div>
    <div class='mc mc-red'><div class='mc-num' style='color:#F87171'>{elen}</div>
      <div class='mc-lbl'>Elendi</div></div>
    </div>""", unsafe_allow_html=True)

    arama_col3, bos3 = st.columns([1.5, 4.5])
    with arama_col3:
        arama_geri = st.text_input("\U0001f50d Hisse Ara", placeholder="Kod ara...",
                                    label_visibility="collapsed", key="geri_ara_v2")

    c1,c2,c3,c4 = st.columns([1.5,2,1.5,1.5])
    with c1:
        kf = st.multiselect("Karar",["GUCLU ADAY","POTANSIYEL","ZAYIF","ELENDI"],
                             default=["GUCLU ADAY","POTANSIYEL"])
    with c2:
        sf = st.multiselect("Sektor", sorted(set(r['sektor'] for r in sonuclar if r['sektor'])))
    with c3:
        sir = st.selectbox("Sirala",[
            "Puan \u2193","FK/PD% \u2193","FK/PD% \u2191",
            f"EFK {yil}Y% \u2193",f"PD {yil}Y% \u2191","F/K \u2191","PD/DD \u2191"])
    with c4:
        sadece_geride = st.checkbox("\U0001f7e2 Sadece Fiyat Geride",False)

    goster = [r for r in sonuclar if r['karar'] in kf]
    if arama_geri: goster = [r for r in goster if arama_geri.upper() in r['kod'].upper()]
    if sf: goster=[r for r in goster if r['sektor'] in sf]
    if sadece_geride: goster=[r for r in goster if r.get('fiyat_geride')]

    def _sv(r,k): return r.get(k) if r.get(k) is not None else -999999
    if sir=="FK/PD% \u2193": goster.sort(key=lambda r:_sv(r,'fkpd'),reverse=True)
    elif sir=="FK/PD% \u2191": goster.sort(key=lambda r:_sv(r,'fkpd') if r.get('fkpd') else 999999)
    elif sir.startswith("EFK"): goster.sort(key=lambda r:_sv(r,'efk_buy'),reverse=True)
    elif "PD" in sir and "Y%" in sir: goster.sort(key=lambda r:_sv(r,'pd_buy') if r.get('pd_buy') else 999999)
    elif sir=="F/K \u2191": goster.sort(key=lambda r:_sv(r,'fk_oran') if r.get('fk_oran') else 999999)
    elif sir=="PD/DD \u2191": goster.sort(key=lambda r:_sv(r,'pddd') if r.get('pddd') else 999999)

    st.markdown(f"<p style='font-size:11px;color:#475569;margin:8px 0'>"
                f"{len(goster)} hisse · {yil}Y periyot · Donem: <b style='color:#94A3B8'>"
                f"{donem_fmt(st.session_state.son_donem)}</b></p>", unsafe_allow_html=True)

    tablo = pd.DataFrame([{
        '\u2b50':            r['kod'] in st.session_state.watchlist,
        'Kod':               r['kod'],
        'Sektor':            r['sektor'],
        'Puan':              int(r['puan']),
        'Karar':             r['karar'],
        'F/K':               round(r['fk_oran'],1) if r.get('fk_oran') else None,
        'PD/DD':             round(r['pddd'],2) if r.get('pddd') else None,
        'FK/PD%':            round(r['fkpd'],1) if r.get('fkpd') else None,
        f'EFK {yil}Y%':     round(r['efk_buy'],0) if r.get('efk_buy') is not None else None,
        f'PD {yil}Y%':      round(r['pd_buy'],0) if r.get('pd_buy') is not None else None,
        f'Satis {yil}Y%':   round(r['ns_buy'],0) if r.get('ns_buy') is not None else None,
        'Fiyat Geride':      '\u2705' if r.get('fiyat_geride') else '\u274c',
        'ROE%':              round(r['roe'],1) if r.get('roe') else None,
        'ROE 30+D':          roe_istikrar_hesapla(engine.quarters, engine.sorted_donems, r['kod'])[0],
        'Piy.Degeri':        fmt_milyon(r.get('pd_val')),
    } for r in goster])

    edited = st.data_editor(tablo, column_config={
        '\u2b50':           st.column_config.CheckboxColumn('\u2b50',width='small'),
        'Kod':               st.column_config.TextColumn('Kod',width='small'),
        'Sektor':            st.column_config.TextColumn('Sektor',width='medium'),
        'Puan':              st.column_config.NumberColumn('Puan',width='small',format='%d'),
        'Karar':             st.column_config.TextColumn('Karar',width='medium'),
        'F/K':               st.column_config.NumberColumn('F/K',width='small',format='%.1f'),
        'PD/DD':             st.column_config.NumberColumn('PD/DD',width='small',format='%.2f'),
        'FK/PD%':            st.column_config.NumberColumn('FK/PD%',width='small',format='%.1f'),
        f'EFK {yil}Y%':     st.column_config.NumberColumn(f'EFK {yil}Y%',width='small',format='%.0f'),
        f'PD {yil}Y%':      st.column_config.NumberColumn(f'PD {yil}Y%',width='small',format='%.0f'),
        f'Satis {yil}Y%':   st.column_config.NumberColumn(f'Satis {yil}Y%',width='small',format='%.0f'),
        'Fiyat Geride':      st.column_config.TextColumn('Fiyat Geride?',width='small'),
        'ROE%':              st.column_config.NumberColumn('ROE%',width='small',format='%.1f'),
        'ROE 30+D':          st.column_config.NumberColumn('ROE 30+D',width='small',format='%d'),
        'Piy.Degeri':        st.column_config.TextColumn('Piy.Degeri',width='small'),
    }, disabled=[c for c in tablo.columns if c!='\u2b50'],
    hide_index=True, use_container_width=True,
    height=min(40+len(goster)*35,600), key='geri_tbl')

    for i,row in edited.iterrows():
        kod,istek = row['Kod'], row['\u2b50']
        in_wl = kod in st.session_state.watchlist
        if istek and not in_wl:
            r=goster[i]
            st.session_state.watchlist[kod]={
                'puan':r['puan'],'karar':r['karar'],'sektor':r['sektor'],
                'sistem':'GERI','eklenme':datetime.now().strftime('%Y-%m-%d'),
                'donem':st.session_state.son_donem}
            st.toast(f"\u2b50 {kod} eklendi!",icon="\u2705"); st.rerun()
        elif not istek and in_wl:
            del st.session_state.watchlist[kod]; st.rerun()

    if goster:
        xls_geri = tablo.drop(columns=['⭐'], errors='ignore')
        st.download_button(
            "⬇️ Listeyi Excel Olarak indir",
            data=df_to_excel_bytes(xls_geri),
            file_name=f"GERI_{st.session_state.son_donem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 3: KESİŞİM
# ════════════════════════════════════════════════════════════════════════════
elif page == "\U0001f3af Kesisim":
    st.markdown(f"""<div class='ph'>
    <div class='ph-badge' style='background:#0A1A0A;color:#4ADE80;border:1px solid #166534'>
    KES\u0130S\u0130M</div>
    <div class='ph-title'>Her \u0130ki Sistemden Gecen Hisseler</div>
    <div class='ph-sub'>FARK filtresi + GER\u0130 filtresi · Operasyonel saglikli + Fiyat gerisinde · En guclu sinyal · {yil}Y periyot</div>
    </div>""", unsafe_allow_html=True)

    veri_yukle_widget()
    engine = st.session_state.engine
    if not engine:
        bos_ekran('\U0001f3af','Veri yukleyerek kesisim analizini baslat'); st.stop()

    kesisim = engine.kesisim_tara(yil)
    fark_n  = len(engine.fark_tara())
    geri_n  = len(engine.geri_tara(yil))
    geride  = [r for r in kesisim if r.get('fiyat_geride')]

    st.markdown(f"""<div class='mrow'>
    <div class='mc mc-blue'><div class='mc-num' style='color:#38BDF8'>{fark_n}</div>
      <div class='mc-lbl'>FARK Gecti</div></div>
    <div class='mc mc-purple'><div class='mc-num' style='color:#A78BFA'>{geri_n}</div>
      <div class='mc-lbl'>GER\u0130 Gecti</div></div>
    <div class='mc mc-green'><div class='mc-num' style='color:#4ADE80'>{len(kesisim)}</div>
      <div class='mc-lbl'>Kesisim</div></div>
    <div class='mc mc-green'><div class='mc-num' style='color:#4ADE80'>{len(geride)}</div>
      <div class='mc-lbl'>Fiyat Geride</div></div>
    <div class='mc mc-blue'><div class='mc-num' style='color:#38BDF8'>
      {round(len(kesisim)/len(engine.son_data)*100,1) if engine.son_data else 0}%</div>
      <div class='mc-lbl'>Evrenin %si</div></div>
    </div>""", unsafe_allow_html=True)

    if not kesisim:
        bos_ekran('\U0001f914','Kesisim bulunamadi — daha fazla donem yuklemeyi dene')
        st.stop()

    # Filtreler
    c1,c2,c3 = st.columns([2,2,1])
    with c1:
        sf = st.multiselect("Sektor", sorted(set(r['sektor'] for r in kesisim if r['sektor'])))
    with c2:
        sir = st.selectbox("Sirala",["Toplam Puan \u2193","FARK Puan \u2193",
                                      "GERI Puan \u2193","FK/PD% \u2193"])
    with c3:
        sg = st.checkbox("\U0001f7e2 Fiyat Geride",False)

    goster = kesisim[:]
    if sf: goster=[r for r in goster if r['sektor'] in sf]
    if sg: goster=[r for r in goster if r.get('fiyat_geride')]
    if sir=="FARK Puan \u2193": goster.sort(key=lambda r:r['fark_puan'],reverse=True)
    elif sir=="GERI Puan \u2193": goster.sort(key=lambda r:r['geri_puan'],reverse=True)
    elif sir=="FK/PD% \u2193": goster.sort(key=lambda r:r.get('fkpd') or 0,reverse=True)

    st.markdown(f"<p style='font-size:11px;color:#475569;margin:8px 0 16px'>"
                f"{len(goster)} hisse listeleniyor</p>", unsafe_allow_html=True)

    for r in goster:
        fk_clr = KARAR_RENK.get(r['fark_karar'],'#94A3B8')
        gk_clr = KARAR_RENK.get(r['geri_karar'],'#94A3B8')
        geride_txt = ("\U0001f7e2 Fiyat Geride" if r.get('fiyat_geride')
                      else "\U0001f534 Fiyat Onde")
        efk_s = f"+{r['efk_buy']:.0f}%" if r.get('efk_buy') is not None else '-'
        pd_s  = f"+{r['pd_buy']:.0f}%"  if r.get('pd_buy') is not None else '-'
        fkpd_s = f"%{r['fkpd']:.1f}"    if r.get('fkpd') else '-'
        pddd_s = f"{r['pddd']:.2f}"      if r.get('pddd') else '-'
        marj_s = f"{r['marj']:.1f}%"     if r.get('marj') else '-'

        col_card, col_btn = st.columns([6,0.5])
        with col_card:
            st.markdown(f"""<div class='kk'>
            <div style='display:flex;justify-content:space-between;align-items:flex-start'>
              <div>
                <span class='kk-kod'>{r['kod']}</span>
                <span class='kk-sektor' style='margin-left:10px'>{r['sektor']}</span>
              </div>
              <span style='color:#475569;font-size:11px'>{geride_txt}</span>
            </div>
            <div class='kk-scores'>
              <div class='kk-score'>
                <div class='kk-score-val' style='color:{fk_clr}'>{r['fark_puan']:.0f}</div>
                <div class='kk-score-lbl'>FARK Puani</div>
              </div>
              <div class='kk-score'>
                <div class='kk-score-val' style='color:{gk_clr}'>{r['geri_puan']:.0f}</div>
                <div class='kk-score-lbl'>GER\u0130 Puani</div>
              </div>
              <div class='kk-score' style='border:1px solid #1E3448'>
                <div class='kk-score-val' style='color:#E2E8F0'>{r['toplam']:.0f}</div>
                <div class='kk-score-lbl'>Toplam</div>
              </div>
            </div>
            <div class='kk-metrics'>
              <div class='kk-chip'>FK/PD: <b>{fkpd_s}</b></div>
              <div class='kk-chip'>EFK {yil}Y: <b>{efk_s}</b></div>
              <div class='kk-chip'>PD {yil}Y: <b>{pd_s}</b></div>
              <div class='kk-chip'>PD/DD: <b>{pddd_s}</b></div>
              <div class='kk-chip'>Marj: <b>{marj_s}</b></div>
              <div class='kk-chip'>Piy: <b>{fmt_milyon(r.get('pd_val'))}</b></div>
              <div class='kk-chip'>A:<b>{r['A']}</b> B:<b>{r['B']}</b> C:<b>{r['C']}</b> D:<b>{r['D']}</b></div>
            </div>
            </div>""", unsafe_allow_html=True)
        with col_btn:
            in_wl = r['kod'] in st.session_state.watchlist
            if st.button('\u2b50' if in_wl else '\u2606', key=f"k_{r['kod']}"):
                if in_wl:
                    del st.session_state.watchlist[r['kod']]
                else:
                    st.session_state.watchlist[r['kod']] = {
                        'puan':r['toplam'],'karar':r['fark_karar'],
                        'sektor':r['sektor'],'sistem':'KESISIM',
                        'fark_puan':r['fark_puan'],'geri_puan':r['geri_puan'],
                        'eklenme':datetime.now().strftime('%Y-%m-%d'),
                        'donem':st.session_state.son_donem}
                    st.toast(f"\u2b50 {r['kod']} eklendi!",icon="\u2705")
                st.rerun()



# ════════════════════════════════════════════════════════════════════════════
# SAYFA: BEBEK HiSSE TARAYICISI
# ════════════════════════════════════════════════════════════════════════════
elif page == "\U0001f476 Bebek Hisse":
    import plotly.graph_objects as go
    st.markdown("""<div class='ph'>
    <div class='ph-badge' style='background:#0A1C0A;color:#4ADE80;border:1px solid #166534'>
    BEBEK HiSSE</div>
    <div class='ph-title'>Karinca Potansiyeli Olan Hisseler</div>
    <div class='ph-sub'>Kucuk cüsse · Büyük kaldirac · Değerlenmemis varliklar · Min 10x potansiyel</div>
    </div>""", unsafe_allow_html=True)

    with st.expander("\U0001f4d6 Metodoloji", expanded=False):
        st.markdown("""<div style='background:#0A1C0A;border:1px solid #166534;
        border-radius:8px;padding:14px 18px;font-size:12px;color:#94A3B8'>
        <b style='color:#4ADE80'>Bebek Hisse Nedir? (GXSMODUJ)</b><br><br>
        "Mevcutta cussesi kucuk (Ozkaynak Birikimi) ama o cusseye gore kaldiraclari buyuk
        (karlilik oranlari, kiymetli istirak, is kolu, degerlenmemis duran varliklar) hisselerdir.
        Karm karinca gucü olan hisseler diyebiliriz — yeri geldiginde bir karinca agirliginin
        50 katı yuku tasiyabilir. Bir bebek hissenin minimum 10 katlik bir potansiyeli olmalidir."
        <br><br>
        <b style='color:#E2E8F0'>Puanlama Boyutlari:</b><br>
        A(25p) Kucuk Cusse · B(30p) Kaldirac Buyuklugu · C(20p) Degerlenmemis Varlik ·
        D(15p) Finansal Saglik · E(10p) Buyume Potansiyeli
        </div>""", unsafe_allow_html=True)

    veri_yukle_widget()
    engine = st.session_state.engine
    if not engine:
        bos_ekran("\U0001f476", "Veri yukleyerek Bebek Hisse taramasini baslat"); st.stop()

    sonuclar = engine.bebek_tara()
    guclu = [r for r in sonuclar if r['puan']>=80]
    aday  = [r for r in sonuclar if 60<=r['puan']<80]
    izle  = [r for r in sonuclar if 40<=r['puan']<60]
    on_x  = [r for r in sonuclar if r.get('potansiyel_x') and r['potansiyel_x']>=10]

    st.markdown(f"""<div class='mrow'>
    <div class='mc mc-green'><div class='mc-num' style='color:#4ADE80'>{len(guclu)}</div>
      <div class='mc-lbl'>Karinca Guclu</div></div>
    <div class='mc mc-yellow'><div class='mc-num' style='color:#FCD34D'>{len(aday)}</div>
      <div class='mc-lbl'>Bebek Aday</div></div>
    <div class='mc mc-blue'><div class='mc-num' style='color:#38BDF8'>{len(izle)}</div>
      <div class='mc-lbl'>Izle</div></div>
    <div class='mc mc-green'><div class='mc-num' style='color:#4ADE80'>{len(on_x)}</div>
      <div class='mc-lbl'>10x+ Potansiyel</div></div>
    <div class='mc mc-purple'><div class='mc-num' style='color:#A78BFA'>{len(sonuclar)}</div>
      <div class='mc-lbl'>Toplam Gecti</div></div>
    </div>""", unsafe_allow_html=True)

    c1,c2,c3 = st.columns([1.5,2,1.5])
    with c1:
        kf = st.multiselect("Karar",["KARINCA GUCLU","BEBEK ADAY","IZLE","ZAYIF"],
                             default=["KARINCA GUCLU","BEBEK ADAY","IZLE"])
    with c2:
        sf = st.multiselect("Sektor", sorted(set(r['sektor'] for r in sonuclar if r['sektor'])))
    with c3:
        sir = st.selectbox("Sirala",["Puan \u2193","Potansiyel X \u2193",
                                      "FK/PD% \u2193","PD/DD \u2191","Marj% \u2193"])
    c4a, c4b = st.columns([1.5, 4.5])
    with c4a:
        min_x = st.number_input("Min Potansiyel X", 0.0, 50.0, 0.0, 1.0)
    with c4b:
        arama = st.text_input("\U0001f50d Hisse Ara", placeholder="Kod: HDFGS, GLRYH...",
                               label_visibility="visible")
    sadece_donus = st.checkbox("🔄 Sadece ROE Donusu Yasayanlar", value=False,
                               help="Son 1-2 donemde ROE negatiften pozitife gecmis hisseler")

    goster = [r for r in sonuclar if r['karar'] in kf]
    if arama: goster = [r for r in goster if arama.upper() in r['kod'].upper()]
    if sf: goster = [r for r in goster if r['sektor'] in sf]
    if min_x > 0: goster = [r for r in goster if r.get('potansiyel_x') and r['potansiyel_x']>=min_x]
    if sadece_donus:
        filtreli = []
        for r in goster:
            donus, _, _ = roe_donus_hesapla(engine.quarters, engine.sorted_donems, r['kod'])
            if donus: filtreli.append(r)
        goster = filtreli

    if sir=="Potansiyel X \u2193":
        goster.sort(key=lambda r: r.get('potansiyel_x') or 0, reverse=True)
    elif sir=="FK/PD% \u2193":
        goster.sort(key=lambda r: r.get('fkpd') or 0, reverse=True)
    elif sir=="PD/DD \u2191":
        goster.sort(key=lambda r: r.get('pddd') or 999)
    elif sir=="Marj% \u2193":
        goster.sort(key=lambda r: r.get('marj') or 0, reverse=True)

    dl1b, dl2b = st.columns([4,1])
    with dl1b:
        st.markdown(f"<p style='font-size:11px;color:#475569;margin:8px 0'>"
                    f"{len(goster)} hisse · Donem: <b style='color:#94A3B8'>"
                    f"{donem_fmt(st.session_state.son_donem)}</b></p>", unsafe_allow_html=True)
    with dl2b:
        if goster:
            import pandas as pd
            bebek_df = pd.DataFrame([{
                'Kod':r['kod'],'Sektor':r['sektor'],'Puan':r['puan'],'Karar':r['karar'],
                'FK/PD%':round(r['fkpd'],1) if r.get('fkpd') else None,
                'PD/DD':round(r['pddd'],2) if r.get('pddd') else None,
                'Marj%':round(r['marj'],1) if r.get('marj') else None,
                'ROE%':round(r['roe'],1) if r.get('roe') else None,
                'Piotroski':r.get('piotroski'),
                'Potansiyel X':round(r['potansiyel_x'],1) if r.get('potansiyel_x') else None,
            } for r in goster])
            st.download_button("⬇ Excel", df_to_excel_bytes(bebek_df),
                               file_name=f"BEBEK_{st.session_state.son_donem}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)

    KARAR_CLR = {'KARINCA GUCLU':'#4ADE80','BEBEK ADAY':'#FCD34D',
                 'IZLE':'#38BDF8','ZAYIF':'#F87171'}

    tablo = pd.DataFrame([{
        '\u2b50':        r['kod'] in st.session_state.watchlist,
        'Kod':           r['kod'],
        'Sektor':        r['sektor'],
        'Puan':          int(r['puan']),
        'Karar':         r['karar'],
        'A':r['A'],'B':r['B'],'C':r['C'],'D':r['D'],'E':r['E'],
        'FK/PD%':        round(r['fkpd'],1)     if r.get('fkpd')      else None,
        'PD/DD':         round(r['pddd'],2)     if r.get('pddd')      else None,
        'Marj%':         round(r['marj'],1)     if r.get('marj')      else None,
        'ROE%':          round(r['roe'],1)      if r.get('roe')       else None,
        'Piotroski':     round(r['piotroski'],0) if r.get('piotroski') else None,
        'Potansiyel X':  round(r['potansiyel_x'],1) if r.get('potansiyel_x') else None,
        'Hedef PD':      fmt_milyon(r.get('hedef_pd')),
        'Mevcut PD':     fmt_milyon(r.get('pd_val')),
        'Ozkaynak':      fmt_milyon(r.get('ozkaynak')),
        'ROE 30+D':      roe_istikrar_hesapla(engine.quarters, engine.sorted_donems, r['kod'])[0],
        'ROE Don':       '🔄' if roe_donus_hesapla(engine.quarters, engine.sorted_donems, r['kod'])[0] else '',
    } for r in goster])

    edited = st.data_editor(tablo, column_config={
        '\u2b50':        st.column_config.CheckboxColumn('\u2b50',width='small'),
        'Kod':            st.column_config.TextColumn('Kod',width='small'),
        'Sektor':         st.column_config.TextColumn('Sektor',width='medium'),
        'Puan':           st.column_config.NumberColumn('Puan',width='small',format='%d'),
        'Karar':          st.column_config.TextColumn('Karar',width='medium'),
        'A':st.column_config.NumberColumn('A',width='small'),
        'B':st.column_config.NumberColumn('B',width='small'),
        'C':st.column_config.NumberColumn('C',width='small'),
        'D':st.column_config.NumberColumn('D',width='small'),
        'E':st.column_config.NumberColumn('E',width='small'),
        'FK/PD%':         st.column_config.NumberColumn('FK/PD%',width='small',format='%.1f'),
        'PD/DD':          st.column_config.NumberColumn('PD/DD',width='small',format='%.2f'),
        'Marj%':          st.column_config.NumberColumn('Marj%',width='small',format='%.1f'),
        'ROE%':           st.column_config.NumberColumn('ROE%',width='small',format='%.1f'),
        'Piotroski':      st.column_config.NumberColumn('Piotroski',width='small',format='%.0f'),
        'Potansiyel X':   st.column_config.NumberColumn('Potansiyel X',width='small',format='%.1f'),
        'Hedef PD':       st.column_config.TextColumn('Hedef PD',width='small'),
        'Mevcut PD':      st.column_config.TextColumn('Mevcut PD',width='small'),
        'Ozkaynak':       st.column_config.TextColumn('Ozkaynak',width='small'),
    }, disabled=[c for c in tablo.columns if c!='\u2b50'],
    hide_index=True, use_container_width=True,
    height=min(40+len(goster)*35,600), key='bebek_tbl')

    for i,row in edited.iterrows():
        kod,istek = row['Kod'], row['\u2b50']
        in_wl = kod in st.session_state.watchlist
        if istek and not in_wl:
            r = goster[i]
            st.session_state.watchlist[kod] = {
                'puan':r['puan'],'karar':r['karar'],'sektor':r['sektor'],
                'sistem':'BEBEK','potansiyel_x':r.get('potansiyel_x'),
                'eklenme':datetime.now().strftime('%Y-%m-%d'),
                'donem':st.session_state.son_donem}
            st.toast(f"\u2b50 {kod} eklendi!",icon="\u2705"); st.rerun()
        elif not istek and in_wl:
            del st.session_state.watchlist[kod]; st.rerun()

    # Potansiyel X Bar Grafigi
    if goster:
        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown("<h3 style=\"color:#E2E8F0;font-size:15px;margin-bottom:10px\">"
                    "\U0001f4ca 10x Potansiyel Haritas\u0131</h3>", unsafe_allow_html=True)
        top20 = sorted([r for r in goster if r.get('potansiyel_x')],
                       key=lambda x: x['potansiyel_x'], reverse=True)[:20]
        if top20:
            kodlar = [r['kod'] for r in top20]
            pot_x  = [r['potansiyel_x'] for r in top20]
            renkler = ['#4ADE80' if x>=10 else '#FCD34D' if x>=5 else '#F87171' for x in pot_x]
            fig = go.Figure(go.Bar(
                x=pot_x, y=kodlar, orientation='h',
                marker_color=renkler,
                text=[f"{x:.1f}x" for x in pot_x],
                textposition='outside',
                textfont=dict(color='#94A3B8',size=11),
                hovertemplate="%{y}: %{x:.1f}x potansiyel<extra></extra>"
            ))
            fig.add_vline(x=10, line_dash="dot", line_color="#4ADE80",
                          annotation_text="10x Esigi", annotation_font_color="#4ADE80")
            fig.update_layout(
                paper_bgcolor='#080E17', plot_bgcolor='#080E17',
                font=dict(color='#475569',size=11),
                margin=dict(l=10,r=60,t=10,b=10),
                height=max(250,len(top20)*28),
                xaxis=dict(gridcolor='#0F2040',title='Potansiyel (X)'),
                yaxis=dict(gridcolor='#0F2040',autorange='reversed'),
                showlegend=False
            )
            st.plotly_chart(fig, use_container_width=True)
            st.markdown("""<div style='background:#0A1C0A;border:1px solid #166534;
            border-radius:8px;padding:10px 16px;font-size:11px;color:#475569;margin-top:4px'>
            \U0001f4a1 <b style='color:#4ADE80'>Potansiyel X Hesabi:</b>
            Teorik Hedef PD = min(EFK x 15, Ozkaynak x 3) / Mevcut PD
            Konservatif tahmindi — spekülasyon degil, finansal temel bazli.
            </div>""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════════════
# SAYFA: DETAY ANALiZi
# ════════════════════════════════════════════════════════════════════════════
elif page == "\U0001f4ca Detay Analizi":
    import plotly.graph_objects as go
    st.markdown("""<div class='ph'>
    <div class='ph-badge' style='background:#0A0A28;color:#A78BFA;border:1px solid #2D1F6E'>
    DETAY ANALiZ</div>
    <div class='ph-title'>Derinlemesine Deger Analizi</div>
    <div class='ph-sub'>Hisseyi sektoru ile karsilastir · Trend grafikleri · Degerleme pozisyonu</div>
    </div>""", unsafe_allow_html=True)

    veri_yukle_widget()
    engine = st.session_state.engine
    if not engine:
        bos_ekran("\U0001f4ca","Once veri yukle"); st.stop()

    yil = st.session_state.geri_yil

    # Hisse secimi — kesisim + FARK + GERI hepsinden
    kesisim   = [r["kod"] for r in engine.kesisim_tara(yil)]
    fark_list = [r["kod"] for r in engine.fark_tara()]
    geri_list = [r["kod"] for r in engine.geri_tara(yil)]
    bebek_kodlar = [r['kod'] for r in engine.bebek_tara() if r['puan']>=50]
    # Oncelik: Kesisim > Bebek > FARK > GERI — kendi icinde alfabetik
    def kategori_sirasi(k):
        if k in kesisim: return (0, k)
        if k in bebek_kodlar: return (1, k)
        if k in fark_list: return (2, k)
        return (3, k)
    tum_liste = sorted(set(kesisim + fark_list + geri_list + bebek_kodlar), key=kategori_sirasi)

    if not tum_liste and not bebek_kodlar:
        st.warning("Analiz icin once tarama yapilmali — veri yukle."); st.stop()

    # Ana liste: Kesisim > FARK > GERI
    ana_liste = sorted(set(kesisim+fark_list+geri_list),
                       key=lambda k: (0 if k in kesisim else 1 if k in fark_list else 2, k))
    bebek_liste = sorted(bebek_kodlar)

    def ana_etiket(k):
        if k in kesisim:   return f"\U0001f3af {k}  \u2190 Kesisim"
        if k in fark_list: return f"\U0001f50d {k}  \u2190 FARK"
        return f"\U0001f4c9 {k}  \u2190 GERI"

    def bebek_et(k):
        if k == "(Sec...)": return "\U0001f476 Bebek hisse sec..."
        extra = []
        if k in kesisim:   extra.append("Kesisim")
        if k in fark_list: extra.append("FARK")
        if k in geri_list: extra.append("GERI")
        suf = " + ".join(extra)
        return f"\U0001f476 {k}" + (f"  \u2190 {suf}" if suf else "")

    col_a, col_b = st.columns(2)
    with col_a:
        v_ana = kesisim[0] if kesisim else (ana_liste[0] if ana_liste else "")
        secilen_ana = st.selectbox(
            "Hisse Sec (Kesisim / FARK / GERI)",
            ana_liste if ana_liste else [""],
            index=ana_liste.index(v_ana) if v_ana in ana_liste else 0,
            format_func=ana_etiket
        )
    with col_b:
        secilen_bebek = st.selectbox(
            "\U0001f476 Bebek Hisse Sec",
            ["(Sec...)"] + bebek_liste,
            format_func=bebek_et
        )

    secilen = secilen_bebek if secilen_bebek != "(Sec...)" else secilen_ana
    sektor = engine.son_data.get(secilen, {}).get("Hisse Sekt\u00f6r", "")
    sektor_hisse_say = sum(1 for v in engine.son_data.values()
                           if v.get("Hisse Sekt\u00f6r","") == sektor)
    st.markdown(
        "<div style='background:#0D1926;border:1px solid #0F2040;border-radius:8px;"
        "padding:10px 16px;margin-top:6px;display:flex;gap:20px;align-items:center'>"
        "<div><div style='font-size:9px;color:#475569;text-transform:uppercase;"
        "letter-spacing:1px'>Secilen Hisse</div>"
        "<div style='font-size:20px;font-weight:900;color:#38BDF8'>" + secilen + "</div></div>"
        "<div><div style='font-size:9px;color:#475569;text-transform:uppercase;"
        "letter-spacing:1px'>Sektor</div>"
        "<div style='font-size:13px;font-weight:700;color:#E2E8F0'>" + sektor + "</div>"
        "<div style='font-size:9px;color:#475569'>" + str(sektor_hisse_say) + " hisse ile karsilastiriliyor</div>"
        "</div></div>",
        unsafe_allow_html=True
    )

    da = DerinAnaliz(engine, secilen)
    deger_kart = da.deger_kart()

    st.markdown("<hr>", unsafe_allow_html=True)
    toplam_donem = len(engine.sorted_donems)

    # ── Degerleme Tablosu — HTML flex ile yatay ──────────────────────────────
    st.markdown("<h3 style=\"color:#E2E8F0;font-size:15px;margin-bottom:10px\">"
                "\U0001f4ca Degerleme Pozisyonu</h3>", unsafe_allow_html=True)

    METRIK_ACIK = {
        'FK/PD%':   'EFK/PD - yuksek=ucuz',
        'PD/DD':    'Dusuk=deger alti',
        'F/K':      'Dusuk=ucuz',
        'ROE%':     'Yuksek=verimli',
        'Marj%':    'Yuksek=kaliteli',
        'Piotroski':'7+=saglikli',
    }

    kartlar_html = "<div style='display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap'>"
    for m in deger_kart:
        if m["durum"] == "veri_yok":
            renk, bg, brd, ico = "#475569", "#0D1926", "#1E3448", ""
        elif m["durum"] == "iyi":
            renk, bg, brd, ico = "#4ADE80", "#071A0F", "#166534", "\u2191"
        else:
            renk, bg, brd, ico = "#F87171", "#1A0707", "#7F1D1D", "\u2193"

        hisse_fmt  = f"{m['hisse']:.1f}"  if m["hisse"]  is not None else "-"
        sektor_fmt = f"{m['sektor']:.1f}" if m["sektor"] is not None else "-"
        fark_fmt   = f"{ico}{m['fark_pct']:+.0f}%" if m["fark_pct"] is not None else ico
        tip = METRIK_ACIK.get(m["isim"], "")

        kartlar_html += (
            f"<div style='flex:1;min-width:85px;background:{bg};border:1px solid {brd};"
            f"border-radius:10px;padding:12px 8px;text-align:center'>"
            f"<div style='font-size:9px;color:#64748B;text-transform:uppercase;"
            f"letter-spacing:1px;margin-bottom:6px;font-weight:600'>{m['isim']}</div>"
            f"<div style='font-size:22px;font-weight:900;color:{renk};line-height:1.1'>{hisse_fmt}</div>"
            f"<div style='font-size:10px;color:{renk};margin:4px 0;font-weight:700'>{fark_fmt}</div>"
            f"<div style='border-top:1px solid {brd};margin:6px 0 4px;padding-top:5px'>"
            f"<div style='font-size:13px;font-weight:700;color:#94A3B8'>{sektor_fmt}</div>"
            f"<div style='font-size:8px;color:#475569;margin-top:1px'>sektor</div></div>"
            f"<div style='font-size:8px;color:#334155;margin-top:3px'>{tip}</div>"
            f"</div>"
        )
    kartlar_html += "</div>"
    st.markdown(kartlar_html, unsafe_allow_html=True)

    # Kac metrikte iyi
    iyi_say = sum(1 for m in deger_kart if m["durum"]=="iyi")
    toplam_say = sum(1 for m in deger_kart if m["durum"]!="veri_yok")
    if toplam_say > 0:
        oran = iyi_say/toplam_say
        renk_oz = "#4ADE80" if oran>=0.6 else "#FCD34D" if oran>=0.4 else "#F87171"
        st.markdown(f"""<div style='background:#0D1926;border:1px solid #0F2040;
        border-radius:8px;padding:10px 16px;margin-top:10px;display:flex;
        align-items:center;gap:12px'>
        <span style='font-size:22px;font-weight:900;color:{renk_oz}'>{iyi_say}/{toplam_say}</span>
        <span style='color:#64748B;font-size:12px'>metrikte sektorunun {"\u00fc" if iyi_say>1 else ""}stunde
        · {"Deger acisindan UCUZ gorunuyor" if oran>=0.6 else "Karma tablo" if oran>=0.4 else "Sektorune gore pahali"}</span>
        </div>""", unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)

    # ── Trend Grafikleri ─────────────────────────────────────────────────────
    st.markdown("<h3 style=\"color:#E2E8F0;font-size:16px;margin-bottom:16px\">"
                "\U0001f4c8 Tarihsel Trend Grafikleri</h3>", unsafe_allow_html=True)

    GRAFIK_RENK_HISSE  = "#38BDF8"
    GRAFIK_RENK_SEKTOR = "#475569"
    GRAFIK_BG          = "#080E17"
    GRAFIK_GRID        = "#0F2040"

    def cizgi_grafik(baslik, hisse_seri, sektor_seri=None, birim=""):
        donems  = [s["donem"] for s in hisse_seri]
        hisse_y = [s["deger"] for s in hisse_seri]

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=donems, y=hisse_y, name=secilen,
            line=dict(color=GRAFIK_RENK_HISSE, width=2.5),
            mode="lines+markers",
            marker=dict(size=5, color=GRAFIK_RENK_HISSE),
            hovertemplate=f"%{{x}}<br>{secilen}: %{{y:.1f}}{birim}<extra></extra>"
        ))
        if sektor_seri:
            sektor_y = [s["deger"] for s in sektor_seri]
            fig.add_trace(go.Scatter(
                x=donems, y=sektor_y, name="Sektor Medyani",
                line=dict(color=GRAFIK_RENK_SEKTOR, width=1.5, dash="dot"),
                mode="lines",
                hovertemplate=f"%{{x}}<br>Sektor: %{{y:.1f}}{birim}<extra></extra>"
            ))
        fig.update_layout(
            title=dict(text=baslik, font=dict(color="#94A3B8", size=13)),
            paper_bgcolor=GRAFIK_BG, plot_bgcolor=GRAFIK_BG,
            font=dict(color="#475569", size=11),
            legend=dict(bgcolor="rgba(0,0,0,0)", font=dict(color="#64748B")),
            margin=dict(l=10, r=10, t=40, b=10),
            height=240,
            xaxis=dict(gridcolor=GRAFIK_GRID, tickangle=-45, tickfont=dict(size=9)),
            yaxis=dict(gridcolor=GRAFIK_GRID),
            hovermode="x unified"
        )
        return fig

    def donem_slider(key, label=""):
        n = st.slider(label or "Donem sayisi", 
                      min_value=1, max_value=toplam_donem, 
                      value=min(toplam_donem, 12), step=1, key=key)
        return [f"{d[:4]}/{d[4:]}" for d in engine.sorted_donems[-n:]]

    # EFK vs Sektor
    efk_h_tum = da.hisse_seri("Esas Faaliyet Kar\u0131 /Zarar\u0131 Net (Y\u0131ll\u0131k)")
    efk_s_tum = da.sektor_seri("Esas Faaliyet Kar\u0131 /Zarar\u0131 Net (Y\u0131ll\u0131k)")

    col1, col2 = st.columns(2)
    with col1:
        n1 = st.slider("", 1, toplam_donem, min(toplam_donem,12), key="sl_efk_v2",
                        format="%d donem", label_visibility="collapsed")
        d1 = [f"{d[:4]}/{d[4:]}" for d in engine.sorted_donems[-n1:]]
        efk_h = [s for s in efk_h_tum if s['donem'] in d1]
        efk_s = [s for s in efk_s_tum if s['donem'] in d1]
        for item in efk_h: item["deger"] = (item["deger"]/1_000_000) if item["deger"] else None
        for item in efk_s: item["deger"] = (item["deger"]/1_000_000) if item["deger"] else None
        st.plotly_chart(cizgi_grafik(
            "Esas Faaliyet Kari (Milyon TL)", efk_h, efk_s, "M"), use_container_width=True)
    with col2:
        n2 = st.slider("", 1, toplam_donem, min(toplam_donem,12), key="sl_pd_v2",
                        format="%d donem", label_visibility="collapsed")
        d2 = [f"{d[:4]}/{d[4:]}" for d in engine.sorted_donems[-n2:]]
        pd_h_tum = da.pd_seri()
        pd_s_tum = da.sektor_seri("Piyasa De\u011feri")
        pd_h = [s for s in pd_h_tum if s['donem'] in d2]
        pd_s = [s for s in pd_s_tum if s['donem'] in d2]
        for item in pd_h: item["deger"] = (item["deger"]/1_000_000) if item["deger"] else None
        for item in pd_s: item["deger"] = (item["deger"]/1_000_000) if item["deger"] else None
        st.plotly_chart(cizgi_grafik(
            "Piyasa Degeri (Milyon TL)", pd_h, pd_s, "M"), use_container_width=True)

    col3, col4 = st.columns(2)
    with col3:
        n3 = st.slider("", 1, toplam_donem, min(toplam_donem,12), key="sl_pddd_v2",
                        format="%d donem", label_visibility="collapsed")
        d3 = [f"{d[:4]}/{d[4:]}" for d in engine.sorted_donems[-n3:]]
        pddd_h = [s for s in da.hisse_seri("Piyasa De\u011feri / Defter De\u011feri") if s['donem'] in d3]
        pddd_s = [s for s in da.sektor_seri("Piyasa De\u011feri / Defter De\u011feri") if s['donem'] in d3]
        st.plotly_chart(cizgi_grafik("PD/DD Trendi", pddd_h, pddd_s), use_container_width=True)
    with col4:
        n4 = st.slider("", 1, toplam_donem, min(toplam_donem,12), key="sl_ns_v2",
                        format="%d donem", label_visibility="collapsed")
        d4 = [f"{d[:4]}/{d[4:]}" for d in engine.sorted_donems[-n4:]]
        ns_h = [s for s in da.hisse_seri("Net Sat\u0131\u015flar (Y\u0131ll\u0131k)") if s['donem'] in d4]
        ns_s = [s for s in da.sektor_seri("Net Sat\u0131\u015flar (Y\u0131ll\u0131k)") if s['donem'] in d4]
        for item in ns_h: item["deger"] = (item["deger"]/1_000_000) if item["deger"] else None
        for item in ns_s: item["deger"] = (item["deger"]/1_000_000) if item["deger"] else None
        st.plotly_chart(cizgi_grafik(
            "Net Satislar (Milyon TL)", ns_h, ns_s, "M"), use_container_width=True)

    # ROE Trend Grafigi
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"color:#E2E8F0;font-size:15px;margin-bottom:8px\">"
                "\U0001f4c8 ROE% Trendi \u2014 %30 Esigi</h3>", unsafe_allow_html=True)
    n_roe = st.slider("", 1, toplam_donem, min(toplam_donem,12), key="sl_roe_v2",
                      format="%d donem", label_visibility="collapsed")
    d_roe = [f"{d[:4]}/{d[4:]}" for d in engine.sorted_donems[-n_roe:]]
    roe_h_tum = da.hisse_seri(C_ROE)
    roe_s_tum = da.sektor_seri(C_ROE)
    roe_h = [s for s in roe_h_tum if s["donem"] in d_roe]
    roe_s = [s for s in roe_s_tum if s["donem"] in d_roe]
    fig_roe = go.Figure()
    fig_roe.add_trace(go.Scatter(
        x=[s["donem"] for s in roe_h], y=[s["deger"] for s in roe_h],
        name=secilen, mode="lines+markers",
        line=dict(color="#38BDF8",width=2), marker=dict(size=5)))
    fig_roe.add_trace(go.Scatter(
        x=[s["donem"] for s in roe_s], y=[s["deger"] for s in roe_s],
        name="Sektor Medyani", mode="lines",
        line=dict(color="#475569",width=1.5,dash="dot")))
    fig_roe.add_hline(y=30, line_color="#F87171", line_dash="dash",
                      annotation_text="%30 Esigi", annotation_font_color="#F87171")
    fig_roe.update_layout(
        paper_bgcolor=GRAFIK_BG, plot_bgcolor=GRAFIK_BG,
        font=dict(color="#475569",size=10),
        margin=dict(l=10,r=10,t=10,b=10), height=280,
        xaxis=dict(gridcolor=GRAFIK_GRID,tickangle=-45,tickfont=dict(size=8)),
        yaxis=dict(gridcolor=GRAFIK_GRID,title="ROE %"),
        legend=dict(font=dict(size=9),bgcolor="rgba(0,0,0,0)"),
        showlegend=True)
    st.plotly_chart(fig_roe, use_container_width=True)
    son_kac, hic_dum, son_roe = roe_istikrar_hesapla(engine.quarters, engine.sorted_donems, secilen)
    if son_roe and son_roe >= 30:
        renk_roe = "#4ADE80" if son_kac >= 8 else "#FCD34D" if son_kac >= 4 else "#38BDF8"
        mesaj_roe = f"Son {son_kac} donemdir %30 uzerinde"
        if hic_dum: mesaj_roe += " | Tum tarihte hic altina dusmedi \u2714"
    else:
        renk_roe = "#F87171"
        mesaj_roe = f"Son ROE: {son_roe:.1f}% \u2014 %30 altinda" if son_roe else "ROE verisi yok"
    st.markdown(
        f"<div style='background:#0D1926;border:1px solid #0F2040;border-radius:8px;"
        f"padding:8px 14px;font-size:11px;color:{renk_roe}'>"
        f"{secilen} \u2014 {mesaj_roe}</div>",
        unsafe_allow_html=True)

    # EFK vs PD Indexlenmis karsilastirma
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"color:#E2E8F0;font-size:16px;margin-bottom:8px\">"
                "\U0001f3af EFK vs PD — Fiyat Geride mi? (Indexlenmis)</h3>",
                unsafe_allow_html=True)
    st.markdown("<p style=\"font-size:11px;color:#475569;margin-bottom:12px\">"
                "Her iki seri ilk doneme gore 100 baz alinarak karsilastirilir. "
                "Mavi cizgi sarin uzegindeyse fiyat geri kalmis demektir.</p>",
                unsafe_allow_html=True)

    n5 = st.slider("", 1, toplam_donem, min(toplam_donem,12), key="sl_idx_v2",
                    format="%d donem", label_visibility="collapsed")
    d5 = [f"{d[:4]}/{d[4:]}" for d in engine.sorted_donems[-n5:]]
    efk_raw = [s for s in da.hisse_seri("Esas Faaliyet Kar\u0131 /Zarar\u0131 Net (Y\u0131ll\u0131k)") if s['donem'] in d5]
    pd_raw  = [s for s in da.pd_seri() if s['donem'] in d5]
    donems  = [s["donem"] for s in efk_raw]

    efk_vals = [s["deger"] for s in efk_raw]
    pd_vals  = [s["deger"] for s in pd_raw]

    efk_base = next((v for v in efk_vals if v and v>0), None)
    pd_base  = next((v for v in pd_vals  if v and v>0), None)

    if efk_base and pd_base:
        efk_idx = [(v/efk_base*100) if v else None for v in efk_vals]
        pd_idx  = [(v/pd_base*100)  if v else None for v in pd_vals]

        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(
            x=donems, y=efk_idx, name="EFK (index)",
            line=dict(color="#38BDF8", width=2.5),
            mode="lines+markers", marker=dict(size=5),
            hovertemplate="%{x}<br>EFK: %{y:.0f}<extra></extra>"
        ))
        fig2.add_trace(go.Scatter(
            x=donems, y=pd_idx, name="Piyasa Degeri (index)",
            line=dict(color="#FCD34D", width=2.5),
            mode="lines+markers", marker=dict(size=5, symbol="diamond"),
            hovertemplate="%{x}<br>PD: %{y:.0f}<extra></extra>"
        ))
        fig2.add_hline(y=100, line_dash="dot", line_color="#1E3448")
        fig2.update_layout(
            paper_bgcolor=GRAFIK_BG, plot_bgcolor=GRAFIK_BG,
            font=dict(color="#475569",size=11),
            legend=dict(bgcolor="rgba(0,0,0,0)",font=dict(color="#64748B")),
            margin=dict(l=10,r=10,t=20,b=10), height=280,
            xaxis=dict(gridcolor=GRAFIK_GRID,tickangle=-45,tickfont=dict(size=9)),
            yaxis=dict(gridcolor=GRAFIK_GRID,title="Baz = 100"),
            hovermode="x unified"
        )
        st.plotly_chart(fig2, use_container_width=True)

        # Son deger yorumu
        last_efk = next((v for v in reversed(efk_idx) if v), None)
        last_pd  = next((v for v in reversed(pd_idx)  if v), None)
        if last_efk and last_pd:
            fark = last_efk - last_pd
            if fark > 20:
                yorum = f"EFK {fark:.0f} puan daha hizli buyudu — fiyat belirgin sekilde geride kalmis."
                yorum_renk = "#4ADE80"
            elif fark > 0:
                yorum = f"EFK {fark:.0f} puan onunde — hafif deger farki var."
                yorum_renk = "#FCD34D"
            else:
                yorum = f"Fiyat {abs(fark):.0f} puan onde — piyasa EFK buyumesini fiyatlamis."
                yorum_renk = "#F87171"
            st.markdown(f"""<div style='background:#0D1926;border-left:3px solid {yorum_renk};
            border-radius:8px;padding:10px 16px;margin-top:4px'>
            <span style='color:{yorum_renk};font-weight:700'>{secilen}</span>
            <span style='color:#94A3B8;font-size:13px'> — {yorum}</span>
            </div>""", unsafe_allow_html=True)
    else:
        st.info("Yeterli veri yok — daha fazla donem yukle.")

    # ── Kat Buyume Tablosu ───────────────────────────────────────────────────
    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown("<h3 style=\"color:#E2E8F0;font-size:16px;margin-bottom:8px\">"
                "\U0001f4ca Kat Buyume Tablosu</h3>", unsafe_allow_html=True)

    kat_c1, kat_c2 = st.columns([3,3])
    with kat_c1:
        st.markdown("<p style=\"font-size:11px;color:#475569;margin:4px 0 8px\">"
                    "Ilk pozitif degerden son degere buyume. GXSMODUJ varlik analizi cercevesi.</p>",
                    unsafe_allow_html=True)
    with kat_c2:
        kat_n = st.slider("", 2, toplam_donem, toplam_donem, key="sl_kat_v2",
                           format="%d donem", label_visibility="collapsed")

    # Engine'i secili donem araligina gore yeniden olustur
    kat_donems = engine.sorted_donems[-kat_n:]
    kat_quarters = {d: engine.quarters[d] for d in kat_donems}
    kat_eng = UnifiedEngine(kat_quarters)
    kat_da = DerinAnaliz(kat_eng, secilen)
    kat_tablo = kat_da.kat_buyume_tablosu()
    col_tbl, col_bar = st.columns([1, 2])

    with col_tbl:
        for item in kat_tablo:
            if item["kat"] is None: continue
            kat = item["kat"]
            if item["kalem"] in ("KV Borc","UV Borc"):
                renk = "#F87171" if kat >= 5 else "#FCD34D" if kat >= 2 else "#4ADE80"
            else:
                renk = "#4ADE80" if kat >= 5 else "#FCD34D" if kat >= 2 else "#F87171"
            st.markdown(f"""<div style='display:flex;justify-content:space-between;
            align-items:center;padding:6px 12px;background:#0D1926;
            border-radius:6px;margin-bottom:3px;border-left:3px solid {renk}'>
            <span style='font-size:12px;color:#94A3B8'>{item["kalem"]}</span>
            <span style='font-size:14px;font-weight:800;color:{renk}'>{kat}x</span>
            </div>""", unsafe_allow_html=True)

    with col_bar:
        kalemler = [i["kalem"] for i in kat_tablo if i["kat"] is not None]
        kat_vals = [i["kat"]   for i in kat_tablo if i["kat"] is not None]
        renkler  = []
        for i in kat_tablo:
            if i["kat"] is None: continue
            if i["kalem"] in ("KV Borc","UV Borc"):
                renkler.append("#F87171" if i["kat"]>=5 else "#FCD34D" if i["kat"]>=2 else "#4ADE80")
            else:
                renkler.append("#4ADE80" if i["kat"]>=5 else "#FCD34D" if i["kat"]>=2 else "#F87171")
        max_val = max(kat_vals) if kat_vals else 1
        text_poz = ["inside" if v > max_val*0.4 else "outside" for v in kat_vals]
        fig_kat = go.Figure(go.Bar(
            x=kat_vals, y=kalemler, orientation="h",
            marker_color=renkler,
            text=[f"{v}x" for v in kat_vals],
            textposition=text_poz,
            insidetextfont=dict(color="#080E17", size=11),
            outsidetextfont=dict(color="#94A3B8", size=11),
            hovertemplate="%{y}: %{x:.1f}x<extra></extra>",
            cliponaxis=False
        ))
        fig_kat.update_layout(
            paper_bgcolor=GRAFIK_BG, plot_bgcolor=GRAFIK_BG,
            font=dict(color="#475569", size=11),
            margin=dict(l=10, r=80, t=10, b=10),
            height=max(200, len(kalemler)*34),
            xaxis=dict(gridcolor=GRAFIK_GRID, title="Kat Buyume",
                       range=[0, max_val*1.2]),
            yaxis=dict(gridcolor=GRAFIK_GRID, autorange="reversed"),
            showlegend=False
        )
        st.plotly_chart(fig_kat, use_container_width=True)

    # EFK ve PD bar grafikleri
    st.markdown("<h3 style=\"color:#E2E8F0;font-size:15px;margin:16px 0 8px\">"
                "EFK ve Piyasa Degeri — Donem Bazli</h3>", unsafe_allow_html=True)
    efk_item = next((i for i in kat_tablo if i["kalem"]=="EFK"), None)
    pd_item  = next((i for i in kat_tablo if i["kalem"]=="Piyasa Degeri"), None)
    if efk_item and pd_item:
        col_e,col_p = st.columns(2)
        with col_e:
            seri = efk_item["seri"]
            fig_e = go.Figure(go.Bar(
                x=[s["donem"] for s in seri],
                y=[(s["deger"]/1_000_000) if s["deger"] else None for s in seri],
                marker_color="#38BDF8",
                hovertemplate="%{x}<br>EFK: %{y:.0f}M<extra></extra>"
            ))
            fig_e.update_layout(
                title=dict(text="EFK (Milyon TL)",font=dict(color="#94A3B8",size=12)),
                paper_bgcolor=GRAFIK_BG,plot_bgcolor=GRAFIK_BG,
                font=dict(color="#475569",size=10),
                margin=dict(l=10,r=10,t=36,b=10),height=220,
                xaxis=dict(gridcolor=GRAFIK_GRID,tickangle=-45,tickfont=dict(size=8)),
                yaxis=dict(gridcolor=GRAFIK_GRID),showlegend=False
            )
            st.plotly_chart(fig_e, use_container_width=True)
        with col_p:
            seri = pd_item["seri"]
            fig_p = go.Figure(go.Bar(
                x=[s["donem"] for s in seri],
                y=[(s["deger"]/1_000_000) if s["deger"] else None for s in seri],
                marker_color="#FCD34D",
                hovertemplate="%{x}<br>PD: %{y:.0f}M<extra></extra>"
            ))
            fig_p.update_layout(
                title=dict(text="Piyasa Degeri (Milyon TL)",font=dict(color="#94A3B8",size=12)),
                paper_bgcolor=GRAFIK_BG,plot_bgcolor=GRAFIK_BG,
                font=dict(color="#475569",size=10),
                margin=dict(l=10,r=10,t=36,b=10),height=220,
                xaxis=dict(gridcolor=GRAFIK_GRID,tickangle=-45,tickfont=dict(size=8)),
                yaxis=dict(gridcolor=GRAFIK_GRID),showlegend=False
            )
            st.plotly_chart(fig_p, use_container_width=True)

    # FAVOK bonus grafigi
    favok_item = next((i for i in kat_tablo if i["kalem"]=="FAVOK"), None)
    if favok_item and any(s["deger"] for s in favok_item["seri"]):
        st.markdown("<h3 style=\"color:#E2E8F0;font-size:15px;margin:8px 0\">FAVOK vs EFK Trendi</h3>",
                    unsafe_allow_html=True)
        seri_f = favok_item["seri"]
        efk_seri_f = da.hisse_seri("Esas Faaliyet Kar\u0131 /Zarar\u0131 Net (Y\u0131ll\u0131k)")
        fig_f = go.Figure()
        fig_f.add_trace(go.Bar(
            x=[s["donem"] for s in seri_f],
            y=[(s["deger"]/1_000_000) if s["deger"] else None for s in seri_f],
            marker_color="#A78BFA",name="FAVOK",
            hovertemplate="%{x}<br>FAVOK: %{y:.0f}M<extra></extra>"
        ))
        fig_f.add_trace(go.Scatter(
            x=[s["donem"] for s in efk_seri_f],
            y=[(s["deger"]/1_000_000) if s["deger"] else None for s in efk_seri_f],
            name="EFK",line=dict(color="#38BDF8",width=2,dash="dot"),
            mode="lines+markers",marker=dict(size=4),
        ))
        fig_f.update_layout(
            paper_bgcolor=GRAFIK_BG,plot_bgcolor=GRAFIK_BG,
            font=dict(color="#475569",size=10),
            legend=dict(bgcolor="rgba(0,0,0,0)",font=dict(color="#64748B")),
            margin=dict(l=10,r=10,t=10,b=10),height=220,
            xaxis=dict(gridcolor=GRAFIK_GRID,tickangle=-45,tickfont=dict(size=8)),
            yaxis=dict(gridcolor=GRAFIK_GRID,title="Milyon TL"),
            hovermode="x unified"
        )
        st.plotly_chart(fig_f, use_container_width=True)


# ════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════
# SAYFA: ROE TARAYICI
# ════════════════════════════════════════════════════════════════════════════
elif page == "\U0001f504 ROE Tarayici":
    import plotly.graph_objects as go
    st.markdown("""<div class='ph'>
    <div class='ph-badge' style='background:#0A1C0A;color:#4ADE80;border:1px solid #166534'>
    ROE TARAYICI</div>
    <div class='ph-title'>ROE Istikrar ve Donus Analizi</div>
    <div class='ph-sub'>Pirlanta hisseler · Donus sinyali · GXSMODUJ metodolojisi</div>
    </div>""", unsafe_allow_html=True)

    veri_yukle_widget()
    engine = st.session_state.engine
    if not engine:
        bos_ekran("\U0001f504", "Once veri yukle"); st.stop()

    son_data = engine.son_data
    donems   = engine.sorted_donems

    # Tum hisseleri tara
    istikrar_list = []  # Hic %30 altina dusmedi
    donus_list    = []  # Son donemde neg->poz
    yaklasan_list = []  # ROE yukseliyor, %30'a yaklasiyor

    for kod, row in son_data.items():
        pd_val = hesapla_pd(row)
        if not pd_val or pd_val <= 0: continue

        son_kac, hic_dum, son_roe = roe_istikrar_hesapla(engine.quarters, donems, kod)
        donus, d_roe, d_kac = roe_donus_hesapla(engine.quarters, donems, kod)

        sektor = row.get("Hisse Sekt\u00f6r", "")

        if hic_dum and son_kac >= 4:
            istikrar_list.append({'kod':kod,'sektor':sektor,'son_roe':son_roe,
                                   'son_kac':son_kac,'pd_val':pd_val})
        if donus:
            donus_list.append({'kod':kod,'sektor':sektor,'son_roe':d_roe,
                                'kac_poz':d_kac,'pd_val':pd_val})

        # ROE yukseliyor ve %20-30 arasi = yaklasanlar
        if son_roe and 15 <= son_roe < 30:
            roe_seri = [safe_float(engine.quarters[d].get(kod,{}).get(C_ROE,""))
                        for d in donems[-6:]]
            gecerli = [r for r in roe_seri if r is not None]
            if len(gecerli) >= 3 and gecerli[-1] > gecerli[0]:
                yaklasan_list.append({'kod':kod,'sektor':sektor,'son_roe':son_roe,
                                       'pd_val':pd_val})

    istikrar_list.sort(key=lambda x: x['son_roe'], reverse=True)
    donus_list.sort(key=lambda x: x['son_roe'], reverse=True)
    yaklasan_list.sort(key=lambda x: x['son_roe'], reverse=True)

    # Ozet metrikler
    st.markdown(f"""<div class='mrow'>
    <div class='mc mc-green'><div class='mc-num' style='color:#4ADE80'>{len(istikrar_list)}</div>
      <div class='mc-lbl'>Pirlanta (Hic Dusmedi)</div></div>
    <div class='mc mc-blue'><div class='mc-num' style='color:#38BDF8'>{len(donus_list)}</div>
      <div class='mc-lbl'>Donus Sinyali</div></div>
    <div class='mc mc-yellow'><div class='mc-num' style='color:#FCD34D'>{len(yaklasan_list)}</div>
      <div class='mc-lbl'>%30'a Yaklasan</div></div>
    </div>""", unsafe_allow_html=True)

    tab_ist, tab_don, tab_yak = st.tabs([
        f"\U0001f48e Pirlanta ({len(istikrar_list)})",
        f"\U0001f504 Donus Sinyali ({len(donus_list)})",
        f"\U0001f4c8 %30'a Yaklasan ({len(yaklasan_list)})"
    ])

    def roe_tablo(liste, ekstra_col=""):
        if not liste:
            st.info("Bu kategoride hisse yok.")
            return
        import pandas as pd
        df = pd.DataFrame([{
            'Kod': r['kod'], 'Sektor': r['sektor'],
            'Son ROE%': round(r['son_roe'],1) if r.get('son_roe') else None,
            ekstra_col if ekstra_col else 'PD': (
                r.get('son_kac') if ekstra_col=='30+ Donem' else
                r.get('kac_poz') if ekstra_col=='Poz. Donem' else None
            ) if ekstra_col else fmt_milyon(r['pd_val']),
            'PD': fmt_milyon(r['pd_val']),
        } for r in liste])
        if not ekstra_col:
            df = df.drop(columns=[""])
        st.dataframe(df, hide_index=True, use_container_width=True,
                     height=min(40+len(liste)*35, 500))
        if liste:
            xls = df.copy()
            st.download_button("\u2b07\ufe0f Excel Indir",
                data=df_to_excel_bytes(xls),
                file_name=f"ROE_tarama_{st.session_state.son_donem}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab_ist:
        st.markdown("""<div style='background:#0A1C0A;border:1px solid #166534;
        border-radius:8px;padding:10px 16px;margin-bottom:10px;font-size:12px;color:#64748B'>
        Tum veri gecmisinde ROE hic %30 altina dusmemis hisseler. GXSMODUJ pirlanta formulu.
        </div>""", unsafe_allow_html=True)
        roe_tablo(istikrar_list, '30+ Donem')

    with tab_don:
        st.markdown("""<div style='background:#0A1020;border:1px solid #1E3A8A;
        border-radius:8px;padding:10px 16px;margin-bottom:10px;font-size:12px;color:#64748B'>
        Son 1-2 donemde ROE negatiften pozitife gecen hisseler. Backtest: ort 22.8x, %9 zarar.
        2018-2020 araliginda bu sinyal SIFIR zarar uretemistir.
        </div>""", unsafe_allow_html=True)
        roe_tablo(donus_list, 'Poz. Donem')

    with tab_yak:
        st.markdown("""<div style='background:#1C1208;border:1px solid #92400E;
        border-radius:8px;padding:10px 16px;margin-bottom:10px;font-size:12px;color:#64748B'>
        ROE yukseliyor ve %15-30 bandinda. Henuz %30 esigine gelmemis, takipte tutulabilir.
        </div>""", unsafe_allow_html=True)
        roe_tablo(yaklasan_list, '')

# SAYFA 4: TAKİP LİSTESİ
# ════════════════════════════════════════════════════════════════════════════
elif page == "\u2b50 Takip Listesi":
    st.markdown("""<div class='ph'>
    <div class='ph-badge' style='background:#1C1208;color:#FCD34D;border:1px solid #92400E'>
    TAKIP</div>
    <div class='ph-title'>Takip Listesi</div>
    <div class='ph-sub'>FARK · GERI · KESISIM sistemlerinden eklenen hisseler</div>
    </div>""", unsafe_allow_html=True)

    wl = st.session_state.watchlist
    if not wl:
        bos_ekran('\u2b50','Herhangi bir sayfadan hisse ekle'); st.stop()

    c1,c2 = st.columns(2)
    with c1:
        st.download_button("\U0001f4be Indir (JSON)",
                            json.dumps(wl,ensure_ascii=False,indent=2),
                            "bist_takip.json","application/json")
    with c2:
        imp = st.file_uploader("JSON Yukle",type=['json'],label_visibility="collapsed")
        if imp:
            try: wl.update(json.loads(imp.read())); st.success("Yuklendi"); st.rerun()
            except: st.error("Format hatali")

    st.markdown("<hr>", unsafe_allow_html=True)

    sistem_gruplar = {}
    for kod,bilgi in wl.items():
        s = bilgi.get('sistem','DIGER')
        sistem_gruplar.setdefault(s,[]).append((kod,bilgi))

    for sistem, items in sistem_gruplar.items():
        renk = {'FARK':'#38BDF8','GERI':'#A78BFA','KESISIM':'#4ADE80'}.get(sistem,'#94A3B8')
        st.markdown(f"<h3 style='color:{renk};font-size:14px;letter-spacing:2px;"
                    f"text-transform:uppercase;margin:16px 0 8px'>{sistem} ({len(items)})</h3>",
                    unsafe_allow_html=True)
        for kod,bilgi in items:
            engine = st.session_state.engine
            yeni_fark = engine.fark_analiz(kod) if engine and kod in engine.son_data else None
            yeni_geri = engine.geri_analiz(kod,yil) if engine and kod in engine.son_data else None

            col_k,col_d,col_rm = st.columns([1.2,4,0.5])
            with col_k:
                st.markdown(f"<b style='color:#38BDF8;font-size:14px'>{kod}</b><br>"
                            f"<span style='color:#475569;font-size:10px'>{bilgi.get('sektor','')}</span>",
                            unsafe_allow_html=True)
            with col_d:
                parts = []
                if yeni_fark:
                    parts.append(f"FARK: <b style='color:#38BDF8'>{yeni_fark['puan']:.0f}p</b>")
                if yeni_geri:
                    efk_s = f"+{yeni_geri['efk_buy']:.0f}%" if yeni_geri.get('efk_buy') is not None else '-'
                    pd_s  = f"+{yeni_geri['pd_buy']:.0f}%"  if yeni_geri.get('pd_buy')  is not None else '-'
                    geride_ico = '\U0001f7e2' if yeni_geri.get('fiyat_geride') else '\U0001f534'
                    parts.append(f"GER\u0130: <b style='color:#A78BFA'>{yeni_geri['puan']:.0f}p</b>")
                    parts.append(f"EFK {yil}Y: <b>{efk_s}</b>")
                    parts.append(f"PD {yil}Y: <b>{pd_s}</b>")
                    parts.append(geride_ico)
                parts.append(f"<span style='color:#1E3448'>{bilgi.get('eklenme','')}</span>")
                st.markdown(f"<span style='font-size:11px;color:#64748B'>" +
                            " · ".join(parts) + "</span>", unsafe_allow_html=True)
            with col_rm:
                if st.button("\U0001f5d1\ufe0f",key=f"rm_{kod}"):
                    del st.session_state.watchlist[kod]; st.rerun()
            st.markdown("<div style='border-bottom:1px solid #080E17;margin:4px 0'></div>",
                        unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 5: METODOLOJİ
# ════════════════════════════════════════════════════════════════════════════
elif page == "\U0001f4da Metodoloji":
    st.markdown("""<div class='ph'>
    <div class='ph-badge' style='background:#0A1628;color:#38BDF8;border:1px solid #1E3448'>
    METODOLOJI</div>
    <div class='ph-title'>Sistem Dokumantasyonu</div>
    <div class='ph-sub'>FARK + GERI + KESISIM metodolojisi · GXSMODUJ uzantisi</div>
    </div>""", unsafe_allow_html=True)

    tab1,tab2,tab3,tab4,tab5 = st.tabs(["FARK","GERI","Bebek Hisse","ROE Istikrari","Kullanim"])

    with tab1:
        st.markdown("""<div style='background:#0D1926;border-left:3px solid #38BDF8;
        border-radius:8px;padding:14px 18px;margin-bottom:12px;font-style:italic;color:#64748B'>
        Operasyonel buyumesi olan ve piyasanin geriden fiyatlanan hisseleri tespit eder.
        </div>""", unsafe_allow_html=True)
        for fno,baslik,aciklama in [
            ("F1","Is Modeli","Holding/GYO/Portfoy eler. FK tutarli operasyonel holding gecer."),
            ("F2","Faal. Kari Devamliligi","Son 8 ceyregin 6+'i pozitif EFK."),
            ("F3","Buyume Varligi","EFK 2Y oncesine gore %20+ buyumeli (ucuzsa %5 esigi)."),
            ("F4","Olumcul Zarar","FK pozitifse NK negativligi TMS 29 etkisi sayilir."),
        ]:
            st.markdown(f"""<div style='background:#0D1926;border:1px solid #0F2040;
            border-radius:8px;padding:12px 16px;margin-bottom:6px'>
            <b style='color:#38BDF8'>{fno}</b>
            <b style='color:#E2E8F0;margin-left:8px'>{baslik}</b>
            <p style='color:#475569;font-size:12px;margin:4px 0 0'>{aciklama}</p>
            </div>""", unsafe_allow_html=True)
        st.markdown("<br>**Puanlama:** A(35p) Buyume · B(48p) Deger · C(25p) Karlilik · D(20p) Model",
                    unsafe_allow_html=True)

    with tab2:
        st.markdown("""<div style='background:#0D1926;border-left:3px solid #A78BFA;
        border-radius:8px;padding:14px 18px;margin-bottom:12px;font-style:italic;color:#64748B'>
        EFK buyumesi fiyatinin onunde olan, deger acisindan ucuz hisseleri tespit eder.
        </div>""", unsafe_allow_html=True)
        for fno,baslik,aciklama in [
            ("F1","F/K < 30","Fiyat/Kazanc 30'un altinda olmali. Negatif F/K eler."),
            ("F2","PD/DD < 5","Asiri spekulatif degerlemeyi eler."),
            ("F3","EFK > 0","Zarar eden sirketleri eler."),
        ]:
            st.markdown(f"""<div style='background:#0D1926;border:1px solid #0F2040;
            border-radius:8px;padding:12px 16px;margin-bottom:6px'>
            <b style='color:#A78BFA'>{fno}</b>
            <b style='color:#E2E8F0;margin-left:8px'>{baslik}</b>
            <p style='color:#475569;font-size:12px;margin:4px 0 0'>{aciklama}</p>
            </div>""", unsafe_allow_html=True)
        st.markdown("<br>**Puanlama:** M1(25p) FK/PD% · M2(30p) EFK Buyume · M3(20p) PD Buyume · M4(15p) Satis Buyume · Bonus(10p)",
                    unsafe_allow_html=True)

    with tab3:
        st.markdown("""<div style='background:#0A1C0A;border:1px solid #166534;
        border-radius:10px;padding:14px 18px;margin-bottom:10px'>
        <b style='color:#4ADE80'>GXSMODUJ Tanimi</b>
        <p style='color:#94A3B8;font-size:12px;margin:6px 0 0'>Mevcutta cussesi kucuk ama
        kaldiraclari buyuk hisselerdir. Minimum 10 katlik potansiyeli olmalidir.</p>
        </div>""", unsafe_allow_html=True)
        for hrf,baslik,aciklama in [
            ("A","Kucuk Cusse (25p)","Ozkaynak < 500M: 25p · < 2Mr: 18p · < 10Mr: 10p"),
            ("B","Kaldirac (30p)","Yuksek marj + Yuksek ROE + Yuksek FK/PD%"),
            ("C","Degerlenmemis Varlik (20p)","PD/DD < 1 + Maddi duran varlik / PD"),
            ("D","Finansal Saglik (15p)","Piotroski F Skor + Cari Oran + FAVOK/Fin.Gider"),
            ("E","Buyume (10p)","Ozsermaye buyumesi + Aktif buyume + Net nakit"),
        ]:
            st.markdown(f"""<div style='background:#0D1926;border:1px solid #0F2040;
            border-radius:8px;padding:10px 14px;margin-bottom:5px'>
            <b style='color:#4ADE80'>{hrf}</b>
            <b style='color:#E2E8F0;margin-left:8px'>{baslik}</b>
            <p style='color:#475569;font-size:12px;margin:3px 0 0'>{aciklama}</p>
            </div>""", unsafe_allow_html=True)
        st.markdown("""<div style='background:#0D1926;border:1px solid #FCD34D;
        border-radius:8px;padding:10px 14px;margin-top:8px'>
        <b style='color:#FCD34D'>10x Potansiyel</b>
        <p style='color:#475569;font-size:12px;margin:4px 0 0'>
        Hedef PD = min(EFK x 15, Ozkaynak x 3) / Mevcut PD.
        Backtesting ortalaması 23.6x getiri, %4 zarar.</p>
        </div>""", unsafe_allow_html=True)

    with tab4:
        st.markdown("""<div style='background:#0A1C0A;border:1px solid #166534;
        border-radius:10px;padding:14px 18px;margin-bottom:10px'>
        <b style='color:#4ADE80'>GXSMODUJ Pirlanta Formulu</b>
        <p style='color:#94A3B8;font-size:12px;margin:6px 0 0'>Surdurulebilir buyume orani ve ROE
        istikrari analizi yapanin biri bin olur.</p>
        </div>""", unsafe_allow_html=True)
        for durum,gosterge,aksiyon,renk in [
            ("ROE %30'u ilk kez geci","ROE 30+D = 1-2","Erken giris sinyali","#38BDF8"),
            ("ROE %30 istikrarli","ROE 30+D >= 4","Guclu sinyal","#4ADE80"),
            ("ROE hic %30 altina dusmedi","Tum tarih > %30","Pirlanta — Tut!","#FCD34D"),
            ("ROE %30 altina dustü","Trend kirildi","Dikkatli ol","#F87171"),
        ]:
            st.markdown(f"""<div style='background:#0D1926;border-left:3px solid {renk};
            border-radius:6px;padding:10px 14px;margin-bottom:6px;display:flex;gap:12px'>
            <div style='flex:1'><b style='color:{renk}'>{durum}</b>
            <br><span style='color:#475569;font-size:11px'>{gosterge}</span></div>
            <div style='color:#E2E8F0;font-size:11px;text-align:right'>{aksiyon}</div>
            </div>""", unsafe_allow_html=True)
        st.markdown("""<div style='background:#0D1926;border:1px solid #0F2040;
        border-radius:8px;padding:12px 16px;margin-top:8px'>
        <b style='color:#E2E8F0'>ROE 30+D Kolonu</b>
        <p style='color:#475569;font-size:12px;margin:4px 0 0'>FARK ve GERI tablolarinda
        son kac donem ust uste ROE %30 ustunde oldugu gosterilir.
        Detay Analizi ROE grafiginde %30 esigi cizgisiyle trend izlenir.</p>
        </div>""", unsafe_allow_html=True)

    with tab5:
        adimlar = [
            ("1","\U0001f4ca","Fastweb Kartini Ayarla",
             "13 kolonu secip kaydet: EFK, PD, PD/DD, Marj, Bor/OK, Nakit, NK, "
             "Fiyat Kazanc, PD/EFK, Net Satislar, ROE, Kapanis, Ozkaynaklar"),
            ("2","\U0001f4e5","Her Ceyrek Indir",
             "Spesifik donem sec (Cari Donem degil). Min 8 donem, 5Y icin 20+ donem onerilir."),
            ("3","\U0001f680","Tek Yukle, Uc Sistem",
             "Ayni Excel dosyalari FARK ve GERI icin ayni anda kullanilir. Iki kez yukleme yok."),
            ("4","\U0001f3af","Kesisimi Incele",
             "Her iki sistemden gecen hisseler en guclu sinyali verir. Buradan basla."),
            ("5","\u2b50","Takibe Al","Ilginc buldukların checkbox ile takip listesine ekle."),
        ]
        for no,em,baslik,aciklama in adimlar:
            st.markdown(f"""<div style='display:flex;gap:12px;background:#0D1926;
            border-radius:8px;padding:12px 16px;margin-bottom:8px'>
            <span style='background:#1E40AF;color:white;border-radius:50%;width:26px;height:26px;
            display:flex;align-items:center;justify-content:center;font-weight:800;
            font-size:12px;flex-shrink:0'>{no}</span>
            <div><b style='color:#E2E8F0'>{em} {baslik}</b><br>
            <span style='color:#475569;font-size:12px'>{aciklama}</span></div>
            </div>""", unsafe_allow_html=True)

        st.markdown("""<div style='background:#1C1208;border:1px solid #92400E;
        border-radius:8px;padding:12px 16px;margin-top:12px'>
        <b style='color:#FCD34D'>\u26a0\ufe0f TMS 29</b>
        <span style='color:#94A3B8;font-size:12px'> · Enflasyon muhasebesi EFK ve Net Satisi
        sisdirabilir. Mutlaka coklu donem yukle.</span>
        </div>""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════════════
# SAYFA 6: AYARLAR
# ════════════════════════════════════════════════════════════════════════════
elif page == "\u2699\ufe0f Ayarlar":
    st.markdown("""<div class='ph'>
    <div class='ph-badge' style='background:#0D1926;color:#64748B;border:1px solid #1E3448'>
    AYARLAR</div>
    <div class='ph-title'>Veri ve Sistem Yonetimi</div>
    </div>""", unsafe_allow_html=True)

    c1,c2 = st.columns(2)
    with c1:
        st.markdown("<b style='color:#E2E8F0'>Guncelleme Takvimi</b>", unsafe_allow_html=True)
        for ay,aciklama in [("Mart","2024/12"),("Haziran","2025/03"),
                             ("Eylul","2025/06"),("Aralik","2025/09")]:
            st.markdown(f"""<div style='background:#0D1926;border:1px solid #0F2040;
            border-radius:6px;padding:8px 14px;margin-bottom:4px;
            display:flex;justify-content:space-between'>
            <b style='color:#38BDF8'>{ay}</b>
            <span style='color:#475569;font-size:11px'>{aciklama} bilancolar aciklandi</span>
            </div>""", unsafe_allow_html=True)

        if st.session_state.son_yukleme:
            gun = (datetime.now()-datetime.fromisoformat(st.session_state.son_yukleme)).days
            if gun>85: st.error(f"Veri guncellenmeli — {gun} gun gecti")
            elif gun>60: st.warning(f"Guncelleme yaklasiyor — {gun} gun")
            else: st.success(f"Veri guncel — {gun} gun once yuklendi")

    with c2:
        st.markdown("<b style='color:#E2E8F0'>Mevcut Veri</b>", unsafe_allow_html=True)
        if st.session_state.quarters:
            donems = sorted(st.session_state.quarters.keys())
            engine = st.session_state.engine
            for lbl,val in [
                ("Yuklenen Donem",str(len(donems))),
                ("Ilk Donem",donem_fmt(donems[0])),
                ("Son Donem",donem_fmt(donems[-1])),
                ("Toplam Hisse",str(len(engine.son_data) if engine else '-')),
            ]:
                st.markdown(f"""<div style='background:#0D1926;border:1px solid #0F2040;
                border-radius:6px;padding:8px 14px;margin-bottom:4px;
                display:flex;justify-content:space-between'>
                <span style='color:#475569;font-size:12px'>{lbl}</span>
                <b style='color:#E2E8F0'>{val}</b></div>""", unsafe_allow_html=True)
        else:
            st.markdown("""<div style='background:#0D1926;border:1px dashed #1E3448;
            border-radius:8px;padding:20px;text-align:center;color:#475569'>
            Veri yuklenmedi</div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("\U0001f5d1\ufe0f Tum Veriyi Sifirla", type="secondary"):
            for k in ['quarters','engine','son_donem','son_yukleme']:
                st.session_state[k] = {} if k=='quarters' else None
            st.success("Sifirland\u0131"); st.rerun()

    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown("""<div style='background:#060D18;border:1px solid #0F2040;
    border-radius:8px;padding:16px 20px;font-size:11px;color:#1E3448;font-family:monospace'>
    BIST Analiz Sistemi v2.0<br>
    FARK Sistemi + GERI Tarayici + Kesisim Analizi<br>
    GXSMODUJ Metodolojisi Uzantisi<br><br>
    Tek Excel karti · Tek yukleme · Uc sistem
    </div>""", unsafe_allow_html=True)
